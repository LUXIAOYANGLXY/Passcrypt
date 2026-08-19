# -*- coding: utf-8 -*-
"""
实验三（对齐 PAEE 文件大小实验口径，E2SE 自有阶段）：固定口令长度 16，
测不同文件大小下 Enc / Dec 延迟与 AuthServer 通信量。

不单独记录 PAEE 的 Ext 阶段。

默认：τ = H4(ct, k2)。副本见 file_size_proto_bench_tau_ct2（τ = H4(ct, k2, ct2)）。
明文加密默认 AES-CTR；GCM 对照见 file_size_proto_bench_gcm / --file-cipher gcm。

阶段::
  Enc_key_store  ↔  Give（密钥封装存 AuthServer+S3）
  Enc_file       ↔  本地 AES-CTR 加密文件（**大文件密文不上传**）
  Enc_total      =  Enc_key_store + Enc_file
  Dec_proto      ↔  OPRF + Take（取回 MSK）
  Dec_file       ↔  本地 AES-CTR 解密
  Dec_total      =  Dec_proto + Dec_file

文件大小默认：1, 10, 100, 200, 300, 400, 500 MB。
输出：行=阶段，列=文件大小（mean±std ms）；另记 AuthServer 通信量。

用法::
  python -m experiment.file_size_proto_bench --trials 3 --auto-server
  python -m experiment.file_size_proto_bench --trials 3 --host 54.x.x.x --port 20202 --real-s3
  python -m experiment.file_size_proto_bench --trials 3 --sizes 1 10 100 --auto-server
  # τ 绑定 ct2 副本：
  python -m experiment.file_size_proto_bench_tau_ct2 --trials 3 --auto-server
"""

from __future__ import annotations

import argparse
import contextlib
import csv
import filecmp
import hashlib
import io
import os
import secrets
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from experiment.benchmark_key_proto import (  # noqa: E402
    DEFAULT_HOST,
    DEFAULT_PORT,
    NAME,
    make_client,
    mean_std,
    start_auth_server,
    wait_port,
)

DISPLAY = "E2SE/PBCS(E2EE)"
DEFAULT_SIZES_MB = (1, 10, 100, 200, 300, 400, 500)
PASSWORD_LEN = 16


def make_password(length: int = PASSWORD_LEN) -> str:
    return "a" * length


def write_random_file(path: Path, size_bytes: int, chunk: int = 8 * 1024 * 1024) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    left = size_bytes
    with path.open("wb") as f:
        while left > 0:
            n = min(chunk, left)
            f.write(os.urandom(n))
            left -= n


def file_sha256(path: Path, chunk: int = 8 * 1024 * 1024) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            b = f.read(chunk)
            if not b:
                break
            h.update(b)
    return h.hexdigest()


def run_size_trials(
    *,
    host: str,
    port: int,
    password: str,
    size_mb: int,
    trials: int,
    id_prefix: str,
    work_dir: Path,
    save_ct_dir: Optional[Path],
    real_s3: bool,
    tau_bind_ct2: bool = False,
    file_cipher: str = "ctr",
) -> List[Dict[str, Any]]:
    """
    每个 trial：
      Reg → Give → 本地 Enc(file) → OPRF+Take → 本地 Dec(file)
    大文件密文默认不上传 S3；不单独记录 Ext。

    tau_bind_ct2=True：Give 内先加密 m→ct2，再 τ=H4(ct,k2,ct2)；Take 用 ct2 验签。
    file_cipher：ctr | gcm（明文 m 的对称算法）。
    """
    cipher = (file_cipher or "ctr").strip().lower()
    if cipher not in ("ctr", "gcm"):
        raise ValueError(f"file_cipher must be ctr|gcm, got {file_cipher!r}")
    cipher_label = "AES-GCM" if cipher == "gcm" else "AES-CTR"

    size = size_mb * 1024 * 1024
    plain_path = work_dir / f"plain_{size_mb}mb.bin"
    print(f"[filesize] 生成明文 {size_mb} MB → {plain_path}")
    write_random_file(plain_path, size)
    plain_digest = file_sha256(plain_path)

    rows: List[Dict[str, Any]] = []
    for t in range(1, trials + 1):
        uid = f"{id_prefix}_{size_mb}mb_{t}"
        passphrase = password
        key0 = f"{uid}/sid"
        key1 = f"{uid}/rid"
        ct_path = work_dir / f"{uid}.ct"
        out_path = work_dir / f"{uid}.out"
        tau_note = "τ=H4(ct,k2,ct2)" if tau_bind_ct2 else "τ=H4(ct,k2)"
        print(
            f"[filesize] ---- {size_mb} MB trial {t}/{trials} id={uid} "
            f"{tau_note} m={cipher_label} ----"
        )

        client = make_client(real_s3=real_s3, quiet_mode_print=False)
        bucket = client.bucket_name
        try:
            with contextlib.redirect_stdout(io.StringIO()):
                client.register(host, port, NAME, uid, passphrase, bucket, key0)

                if tau_bind_ct2:
                    # Give 内：Enc(file)→ct2，再 τ=H4(ct,k2,ct2) 并寄存
                    t_give0 = time.perf_counter()
                    msk, give_s, give_r = client.give(
                        host,
                        port,
                        NAME,
                        uid,
                        passphrase,
                        bucket,
                        key1,
                        key0,
                        plaintext_path=str(plain_path),
                        ct2_path=str(ct_path),
                        file_cipher=cipher,
                    )
                    enc_total_ms = (time.perf_counter() - t_give0) * 1000.0
                    enc_file_ms = float(getattr(client, "last_give_file_ms", 0.0) or 0.0)
                    enc_key_ms = max(0.0, enc_total_ms - enc_file_ms)
                else:
                    t_give0 = time.perf_counter()
                    msk, give_s, give_r = client.give(
                        host, port, NAME, uid, passphrase, bucket, key1, key0
                    )
                    enc_key_ms = (time.perf_counter() - t_give0) * 1000.0

                    t_fe0 = time.perf_counter()
                    client.encrypt_file(
                        str(plain_path), str(ct_path), msk, file_cipher=cipher
                    )
                    enc_file_ms = (time.perf_counter() - t_fe0) * 1000.0

                    enc_total_ms = enc_key_ms + enc_file_ms

                # give_s=KS(id,t,ct,τ,e)；give_r=S3(id,pwd,rid,sid)
                enc_comm = int(give_s + give_r)

                if save_ct_dir is not None:
                    save_ct_dir.mkdir(parents=True, exist_ok=True)
                    dest = save_ct_dir / f"{uid}.ct"
                    if ct_path.resolve() != dest.resolve():
                        dest.write_bytes(ct_path.read_bytes())

                t_dec0 = time.perf_counter()
                _, oprf2_s, oprf2_r = client.ib_oprf(host, port, NAME, uid, passphrase)
                take_kw = {}
                if tau_bind_ct2:
                    take_kw["ct2_path"] = str(ct_path)
                mskr, take_s, take_r = client.take(
                    host, port, NAME, uid, passphrase, bucket, key1, key0, **take_kw
                )
                dec_proto_ms = (time.perf_counter() - t_dec0) * 1000.0

            if msk != mskr:
                raise RuntimeError("msk mismatch")

            t_fd0 = time.perf_counter()
            client.decrypt_file(
                str(ct_path), str(out_path), mskr, file_cipher=cipher
            )
            dec_file_ms = (time.perf_counter() - t_fd0) * 1000.0

            if file_sha256(out_path) != plain_digest and not filecmp.cmp(
                plain_path, out_path, shallow=False
            ):
                raise RuntimeError("plaintext mismatch after decrypt")

            dec_total_ms = dec_proto_ms + dec_file_ms
            # OPRF(id,u,v) + Take KS(id,t,ct,τ) + Take S3(id,pwd,rid,sid)
            dec_comm = int(oprf2_s + oprf2_r + take_s + take_r)
            ct_bytes = ct_path.stat().st_size if ct_path.exists() else 0

            row = {
                "success": True,
                "size_mb": size_mb,
                "trial_id": t,
                "id": uid,
                "file_cipher": cipher,
                "enc_key_store_ms": enc_key_ms,
                "enc_file_ms": enc_file_ms,
                "enc_total_ms": enc_total_ms,
                "dec_proto_ms": dec_proto_ms,
                "dec_file_ms": dec_file_ms,
                "dec_total_ms": dec_total_ms,
                "enc_comm_bytes": enc_comm,
                "dec_comm_bytes": dec_comm,
                "ct_bytes": ct_bytes,
                "error": "",
            }
            rows.append(row)
            print(
                f"  Enc={enc_total_ms:.1f}(give={enc_key_ms:.1f}+file={enc_file_ms:.1f}) "
                f"Dec={dec_total_ms:.1f}(proto={dec_proto_ms:.1f}+file={dec_file_ms:.1f}) "
                f"wire_enc={enc_comm}B wire_dec={dec_comm}B ct={ct_bytes}B"
            )
        except Exception as e:  # noqa: BLE001
            print(f"  [fail] {e}")
            rows.append(
                {
                    "success": False,
                    "size_mb": size_mb,
                    "trial_id": t,
                    "id": uid,
                    "file_cipher": cipher,
                    "enc_key_store_ms": 0.0,
                    "enc_file_ms": 0.0,
                    "enc_total_ms": 0.0,
                    "dec_proto_ms": 0.0,
                    "dec_file_ms": 0.0,
                    "dec_total_ms": 0.0,
                    "enc_comm_bytes": 0,
                    "dec_comm_bytes": 0,
                    "ct_bytes": 0,
                    "error": str(e),
                }
            )
        finally:
            for p in (ct_path, out_path):
                try:
                    if p.exists() and save_ct_dir is None:
                        p.unlink()
                except OSError:
                    pass

    try:
        plain_path.unlink()
    except OSError:
        pass
    return rows


PHASE_ROWS = (
    ("Enc（Give+本地加密文件）", "enc_total_ms"),
    ("Dec（OPRF+Take+本地解密文件）", "dec_total_ms"),
)


def breakdown_rows(file_cipher: str = "ctr") -> tuple:
    label = "AES-GCM" if (file_cipher or "ctr").lower() == "gcm" else "AES-CTR"
    return (
        ("Enc_key_store（Give）", "enc_key_store_ms"),
        (f"Enc_file（本地 {label}）", "enc_file_ms"),
        ("Enc_total（Give+file）", "enc_total_ms"),
        ("Dec_proto（OPRF+Take→MSK）", "dec_proto_ms"),
        (f"Dec_file（本地 {label}）", "dec_file_ms"),
        ("Dec_total", "dec_total_ms"),
    )


BREAKDOWN_ROWS = breakdown_rows("ctr")

COMM_ROWS = (
    ("Enc 通信量 KS+S3 (B)", "enc_comm_bytes"),
    ("Dec 通信量 KS+S3 (B)", "dec_comm_bytes"),
)


def summarize_by_size(
    all_raw: List[Dict[str, Any]],
    sizes: Sequence[int],
    *,
    file_cipher: str = "ctr",
) -> Dict[int, Dict[str, Any]]:
    rows_spec = breakdown_rows(file_cipher)
    out: Dict[int, Dict[str, Any]] = {}
    for mb in sizes:
        rows = [r for r in all_raw if r["size_mb"] == mb]
        ok = [r for r in rows if r["success"]]
        stats: Dict[str, Any] = {"n_trials": len(rows), "n_success": len(ok)}
        for _name, key in rows_spec + COMM_ROWS:
            m, s = mean_std([float(r[key]) for r in ok])
            stats[f"{key}_mean"] = m
            stats[f"{key}_std"] = s
        out[mb] = stats
    return out


def _size_header(sizes: Sequence[int]) -> List[str]:
    return ["阶段"] + [f"{mb}MB" for mb in sizes]


def _row_mean_std(
    by_size: Dict[int, Dict[str, Any]],
    sizes: Sequence[int],
    phase: str,
    key: str,
) -> List[Any]:
    cells: List[Any] = [phase]
    for mb in sizes:
        st = by_size[mb]
        cells.append(f"{st[f'{key}_mean']:.2f} ± {st[f'{key}_std']:.2f}")
    return cells


def write_outputs(
    by_size: Dict[int, Dict[str, Any]],
    raw: List[Dict[str, Any]],
    sizes: Sequence[int],
    out_dir: Path,
    *,
    endpoint: str,
    trials: int,
    password_len: int,
    timestamp: str,
    real_s3: bool,
    tau_bind_ct2: bool = False,
    out_stem: str = "file_size_proto_benchmark",
    file_cipher: str = "ctr",
) -> Tuple[Path, Optional[Path]]:
    out_dir.mkdir(parents=True, exist_ok=True)
    cipher = (file_cipher or "ctr").strip().lower()
    cipher_label = "AES-GCM" if cipher == "gcm" else "AES-CTR"
    rows_spec = breakdown_rows(cipher)
    s3_note = "真实 AWS S3（Give/Take 延迟含 S3）" if real_s3 else "内存模拟 S3"
    tau_note = (
        "τ=H4(ct,k2,ct2)（Give 内加密文件并绑定 ct2）"
        if tau_bind_ct2
        else "τ=H4(ct,k2)（不含文件密文）"
    )
    note = (
        f"口令长度={password_len}；trials/大小={trials}；Server={endpoint}；S3={s3_note}；"
        f"明文加密={cipher_label}；{tau_note}；不记录 Ext；"
        "Enc_total=Give+本地加密（大文件密文不上传）；"
        "Dec=OPRF+Take+本地解密；"
        "通信量=论文字段（Give:id,t,ct,τ,e + S3；Dec:OPRF id,u,v + Take id,t,ct,τ + S3）；"
        "列=文件大小"
    )
    header = _size_header(sizes)
    sep = "|" + "|".join(["------"] * len(header)) + "|"

    title = (
        f"# E2SE 文件大小：Enc / Dec（明文 {cipher_label}"
        + ("；τ 绑定 ct2）" if tau_bind_ct2 else "）")
    )
    if not tau_bind_ct2 and cipher == "ctr":
        title = "# E2SE 文件大小：Enc / Dec 延迟（对齐 PAEE 实验三文件大小轴）"

    md = [
        title,
        "",
        f"生成时间: {timestamp}",
        note,
        "",
        "## 主表 mean±std (ms)",
        "",
        "| " + " | ".join(header) + " |",
        sep,
    ]
    for phase, key in PHASE_ROWS:
        md.append(
            "| " + " | ".join(str(c) for c in _row_mean_std(by_size, sizes, phase, key)) + " |"
        )
    md += ["", "## 分解 mean±std (ms)", "", "| " + " | ".join(header) + " |", sep]
    for phase, key in rows_spec:
        md.append(
            "| " + " | ".join(str(c) for c in _row_mean_std(by_size, sizes, phase, key)) + " |"
        )
    md += ["", "## 通信量 KS+S3 字段 mean±std (B)", "", "| " + " | ".join(header) + " |", sep]
    for phase, key in COMM_ROWS:
        md.append(
            "| " + " | ".join(str(c) for c in _row_mean_std(by_size, sizes, phase, key)) + " |"
        )
    md += [
        "",
        "## 阶段说明",
        "",
        f"- Enc = Give + 本地 {cipher_label}；{tau_note}",
        f"- Dec = OPRF + Take + 本地 {cipher_label}",
        "",
    ]
    md_path = out_dir / f"{out_stem}.md"
    md_path.write_text("\n".join(md) + "\n", encoding="utf-8")

    with (out_dir / f"{out_stem}_summary.csv").open(
        "w", newline="", encoding="utf-8"
    ) as f:
        w = csv.writer(f)
        w.writerow(header)
        for phase, key in PHASE_ROWS:
            w.writerow(_row_mean_std(by_size, sizes, phase, key))
        w.writerow([])
        w.writerow(["分解"])
        for phase, key in rows_spec:
            w.writerow(_row_mean_std(by_size, sizes, phase, key))
        w.writerow([])
        w.writerow(["通信量"])
        for phase, key in COMM_ROWS:
            w.writerow(_row_mean_std(by_size, sizes, phase, key))

    if raw:
        with (out_dir / f"{out_stem}.csv").open(
            "w", newline="", encoding="utf-8"
        ) as f:
            w = csv.DictWriter(f, fieldnames=list(raw[0].keys()))
            w.writeheader()
            w.writerows(raw)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        print("[filesize] openpyxl missing; skip xlsx")
        return md_path, None

    def style_h(ws) -> None:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"
    ws.append(header)
    style_h(ws)
    for phase, key in PHASE_ROWS:
        ws.append(_row_mean_std(by_size, sizes, phase, key))
    ws.append([])
    ws.append([note])
    ws.append([f"方案={DISPLAY}", f"时间={timestamp}", f"明文={cipher_label}"])

    ws2 = wb.create_sheet("分解")
    ws2.append(header)
    style_h(ws2)
    for phase, key in rows_spec:
        ws2.append(_row_mean_std(by_size, sizes, phase, key))

    ws3 = wb.create_sheet("延迟_mean")
    ws3.append(header)
    style_h(ws3)
    for phase, key in PHASE_ROWS:
        ws3.append([phase] + [round(by_size[mb][f"{key}_mean"], 3) for mb in sizes])

    ws_c = wb.create_sheet("通信量")
    ws_c.append(header)
    style_h(ws_c)
    for phase, key in COMM_ROWS:
        ws_c.append(_row_mean_std(by_size, sizes, phase, key))

    ws4 = wb.create_sheet("原始数据")
    if raw:
        keys = list(raw[0].keys())
        ws4.append(keys)
        style_h(ws4)
        for r in raw:
            ws4.append([r[k] for k in keys])

    xlsx_path = out_dir / f"{out_stem}.xlsx"
    wb.save(xlsx_path)
    return md_path, xlsx_path


def main(
    argv: Optional[List[str]] = None,
    *,
    tau_bind_ct2: bool = False,
    out_stem: str = "file_size_proto_benchmark",
    id_prefix_base: str = "fsize",
    description: Optional[str] = None,
    file_cipher: str = "ctr",
) -> int:
    parser = argparse.ArgumentParser(
        description=description
        or "E2SE Enc/Dec vs file size（对齐 PAEE 实验三；不记录 Ext）"
    )
    parser.add_argument("--trials", type=int, default=3)
    parser.add_argument(
        "--sizes",
        type=int,
        nargs="+",
        default=list(DEFAULT_SIZES_MB),
        help="文件大小 MB，默认 1 10 100 200 300 400 500",
    )
    parser.add_argument("--password-len", type=int, default=PASSWORD_LEN)
    parser.add_argument("--host", default=None)
    parser.add_argument("--port", type=int, default=None)
    parser.add_argument("--auto-server", action="store_true")
    parser.add_argument(
        "--real-s3",
        action="store_true",
        help="Give/Take 走真实 AWS S3（延迟含 S3；大文件密文仍不上传）",
    )
    parser.add_argument(
        "--save-ct",
        action="store_true",
        help="将文件密文写到 experiment/output/file_cts/（不计时）",
    )
    parser.add_argument(
        "--file-cipher",
        choices=("ctr", "gcm"),
        default=None,
        help="明文 m 加密算法：ctr（默认）或 gcm",
    )
    parser.add_argument("--out-dir", default=str(ROOT / "experiment" / "output"))
    args = parser.parse_args(argv)

    cipher = (args.file_cipher or file_cipher or "ctr").strip().lower()

    password = make_password(args.password_len)
    assert len(password) == args.password_len

    try:
        import Constants

        host = args.host or Constants.AUTH_SERVER_ADDRESS or DEFAULT_HOST
        port = args.port or Constants.AUTH_SERVER_PORT_NUMBER or DEFAULT_PORT
    except Exception:
        host = args.host or DEFAULT_HOST
        port = args.port or DEFAULT_PORT

    if args.host is None and args.auto_server:
        host = DEFAULT_HOST

    proc = None
    try:
        if args.auto_server:
            print(f"[filesize] AuthServer {host}:{port} ...")
            proc = start_auth_server(host, port)
        elif not wait_port(host, port, timeout_s=2.0):
            print(f"[filesize] no server on {host}:{port}; use --auto-server")
            return 1

        make_client(real_s3=args.real_s3, quiet_mode_print=True)
        run_tag = datetime.now().strftime("%Y%m%d%H%M%S") + secrets.token_hex(2)
        id_prefix = f"{id_prefix_base}_{run_tag}"
        out_dir = Path(args.out_dir)
        work_dir = out_dir / "filesize_work" / run_tag
        work_dir.mkdir(parents=True, exist_ok=True)
        save_ct_dir = out_dir / "file_cts" / run_tag if args.save_ct else None

        print(f"[filesize] 明文加密算法={cipher.upper()}")

        all_raw: List[Dict[str, Any]] = []
        for mb in args.sizes:
            print(f"[filesize] ===== file size {mb} MB =====")
            all_raw.extend(
                run_size_trials(
                    host=host,
                    port=port,
                    password=password,
                    size_mb=mb,
                    trials=args.trials,
                    id_prefix=id_prefix,
                    work_dir=work_dir,
                    save_ct_dir=save_ct_dir,
                    real_s3=args.real_s3,
                    tau_bind_ct2=tau_bind_ct2,
                    file_cipher=cipher,
                )
            )

        by_size = summarize_by_size(all_raw, args.sizes, file_cipher=cipher)
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        md, xlsx = write_outputs(
            by_size,
            all_raw,
            args.sizes,
            out_dir,
            endpoint=f"{host}:{port}",
            trials=args.trials,
            password_len=args.password_len,
            timestamp=ts,
            real_s3=args.real_s3,
            tau_bind_ct2=tau_bind_ct2,
            out_stem=out_stem,
            file_cipher=cipher,
        )

        header = _size_header(args.sizes)
        col_w = max(16, max(len(h) for h in header[1:]) + 2)
        print()
        print(f"{header[0]:<42}" + "".join(f"{h:>{col_w}}" for h in header[1:]))
        for phase, key in PHASE_ROWS:
            row = _row_mean_std(by_size, args.sizes, phase, key)
            print(f"{row[0]:<42}" + "".join(f"{c:>{col_w}}" for c in row[1:]))
        print(f"\n报告: {md}")
        if xlsx:
            print(f"Excel: {xlsx}")
        return 0
    finally:
        if proc is not None:
            proc.terminate()
            try:
                proc.wait(timeout=3)
            except Exception:
                proc.kill()
            logf = getattr(proc, "_bench_logf", None)
            if logf is not None:
                logf.close()


if __name__ == "__main__":
    raise SystemExit(main())
