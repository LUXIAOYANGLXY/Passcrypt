# -*- coding: utf-8 -*-
"""
E2SE TCP 协议基准（对标 PAEE Init/Rec；不单独记录 Ext）

阶段::
  Reg(一次性)     ↔  Register（AuthServer；sid 本地/S3 不计 wire）
  Enc_proto≈Init  ↔  OPRF + Give
  Dec_proto≈Rec  ↔  OPRF + Take

延迟：含密码运算 + AuthServer TCP（短连接，建连计入各阶段）。
      默认 sid/rid 用内存模拟；加 ``--real-s3`` 时延迟**含真实 S3** 上下传 RTT。
通信量（论文字段载荷，不含 TCP 帧头/opcode/长度前缀）::
  OPRF: id,u,v；Give: id,t,ct,τ,e；Take: id,t,ct,τ；
  S3: id,pwd,rid,sid（按阶段计入）。
  Enc_proto = OPRF + Give(+S3)；Dec_proto = OPRF + Take(+S3)。
轮数：一对 Client↔Server 请求/响应 = 1 轮（Reg=1, Enc/Dec_proto=2）。

用法::
  # 本机 AuthServer + 内存 S3（协议基线）
  python -m experiment.benchmark_key_proto --trials 20 --warmup 1 --auto-server

  # AuthServer + 真实 AWS S3（延迟含 S3）
  python -m experiment.benchmark_key_proto --trials 20 --host 127.0.0.1 --port 20202 --real-s3
  python -m experiment.benchmark_key_proto --trials 20 --host <EC2_IP> --port 20202 --real-s3
"""

from __future__ import annotations

import argparse
import configparser
import contextlib
import io
import os
import statistics
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DEFAULT_HOST = "127.0.0.1"
DEFAULT_PORT = 20202
NAME = "CN=usyd.authserver,OU=authserver,O=server,L=sydney,ST=NSW,C=AU"
DISPLAY = "E2SE/PBCS(E2EE)"

PHASE_REG = "Reg(一次性)"
PHASE_ENC = "Enc_proto≈Init"
PHASE_DEC = "Dec_proto≈Rec"


class LocalObjectStore:
    def __init__(self) -> None:
        self._obj: Dict[str, bytes] = {}

    def upload_fileobj(self, f, bucket: str, key: str) -> None:
        self._obj[f"{bucket}/{key}"] = f.read()

    def get_object(self, Bucket: str, Key: str):
        return {"Body": io.BytesIO(self._obj[f"{Bucket}/{Key}"])}


def load_aws_config() -> Dict[str, str]:
    cfg = configparser.ConfigParser()
    cfg.read(ROOT / "config.properties", encoding="utf-8")
    return {
        "access_key_id": cfg.get("DEFAULT", "accessKeyId", fallback=""),
        "secret_key_id": cfg.get("DEFAULT", "secretKeyId", fallback=""),
        "region_name": cfg.get("DEFAULT", "regionName", fallback="ap-northeast-3"),
        "bucket_name": cfg.get("DEFAULT", "bucketName", fallback=""),
    }


def mean_std(xs: List[float]) -> Tuple[float, float]:
    if not xs:
        return 0.0, 0.0
    if len(xs) == 1:
        return xs[0], 0.0
    return statistics.mean(xs), statistics.stdev(xs)


def wait_port(host: str, port: int, timeout_s: float = 15.0) -> bool:
    import socket

    deadline = time.time() + timeout_s
    while time.time() < deadline:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(0.5)
        try:
            if sock.connect_ex((host, port)) == 0:
                return True
        except OSError:
            pass
        finally:
            sock.close()
        time.sleep(0.15)
    return False


def start_auth_server(host: str, port: int) -> subprocess.Popen:
    env = os.environ.copy()
    env["PYTHONUNBUFFERED"] = "1"
    log_path = ROOT / "experiment" / "output" / "authserver_bench.log"
    log_path.parent.mkdir(parents=True, exist_ok=True)
    logf = open(log_path, "w", encoding="utf-8")
    proc = subprocess.Popen(
        [sys.executable, "E2seMain.py", "authserver"],
        cwd=str(ROOT),
        stdout=logf,
        stderr=subprocess.STDOUT,
        env=env,
    )
    if not wait_port(host, port):
        proc.terminate()
        logf.close()
        raise RuntimeError(f"AuthServer failed to listen on {host}:{port}; see {log_path}")
    proc._bench_logf = logf  # type: ignore[attr-defined]
    return proc


def make_client(*, real_s3: bool, quiet_mode_print: bool = True) -> Any:
    from Client import Client

    quiet = type("Quiet", (), {"log": staticmethod(lambda *a, **k: None)})()
    if real_s3:
        aws = load_aws_config()
        if not aws["access_key_id"] or not aws["bucket_name"] or aws["access_key_id"] == "local":
            raise RuntimeError(
                "真实 S3 需要 config.properties 中有效的 accessKeyId/secretKeyId/bucketName"
            )
        c = Client(
            aws["access_key_id"],
            aws["secret_key_id"],
            aws["region_name"],
            aws["bucket_name"],
            use_local_sim=False,
        )
        c.verbose = False
        c.logger = quiet
        if quiet_mode_print:
            print(
                f"[bench] S3=AWS real  region={aws['region_name']}  "
                f"bucket={aws['bucket_name']}  （延迟将含 S3 RTT）"
            )
        return c

    c = Client("local", "local", "ap-northeast-3", "local-bucket", use_local_sim=True)
    c.s3_client = LocalObjectStore()
    c.verbose = False
    c.logger = quiet
    if quiet_mode_print:
        print("[bench] S3=内存模拟（延迟不含公网 S3）")
    return c


def run_trial(client: Any, trial_id: int, host: str, port: int) -> Dict[str, float]:
    import binascii
    import random

    rand = random.Random()
    rb = bytes([rand.randint(0, 255) for _ in range(10)])
    user_id = f"bench{binascii.hexlify(rb).decode()}"
    passphrase = f"pass{binascii.hexlify(rb).decode()}"
    key0 = f"{user_id}/sid"
    key1 = f"{user_id}/rid"
    bucket = client.bucket_name

    with contextlib.redirect_stdout(io.StringIO()):
        t0 = time.perf_counter()
        reg_s, reg_r = client.register(host, port, NAME, user_id, passphrase, bucket, key0)
        reg_ms = (time.perf_counter() - t0) * 1000.0
        reg_comm = float(reg_s + reg_r)

        t_enc0 = time.perf_counter()
        t_ext0 = time.perf_counter()
        _, oprf_s, oprf_r = client.ib_oprf(host, port, NAME, user_id, passphrase)
        ext_ms = (time.perf_counter() - t_ext0) * 1000.0
        ext_comm = float(oprf_s + oprf_r)

        msk, give_s, give_r = client.give(
            host, port, NAME, user_id, passphrase, bucket, key1, key0
        )
        enc_proto_ms = (time.perf_counter() - t_enc0) * 1000.0
        enc_proto_comm = ext_comm + float(give_s + give_r)

        t_dec0 = time.perf_counter()
        _, oprf2_s, oprf2_r = client.ib_oprf(host, port, NAME, user_id, passphrase)
        mskr, take_s, take_r = client.take(
            host, port, NAME, user_id, passphrase, bucket, key1, key0
        )
        dec_proto_ms = (time.perf_counter() - t_dec0) * 1000.0
        dec_proto_comm = float(oprf2_s + oprf2_r + take_s + take_r)

    if msk != mskr:
        raise AssertionError(f"trial {trial_id}: MSK mismatch")

    return {
        "reg_ms": reg_ms,
        "reg_comm": reg_comm,
        "ext_ms": ext_ms,
        "ext_comm": ext_comm,
        "enc_proto_ms": enc_proto_ms,
        "enc_proto_comm": enc_proto_comm,
        "dec_proto_ms": dec_proto_ms,
        "dec_proto_comm": dec_proto_comm,
    }


def write_outputs(
    out_dir: Path,
    rows: List[Dict[str, float]],
    trials: int,
    host: str,
    port: int,
    *,
    real_s3: bool,
) -> Tuple[Path, Optional[Path]]:
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    endpoint = f"{host}:{port}"
    s3_mode = "真实 AWS S3（延迟含 S3 RTT）" if real_s3 else "内存模拟 S3（延迟不含公网 S3）"
    note = (
        f"口径：短连接；S3={s3_mode}；"
        "Reg=Register；Enc_proto=OPRF+Give；Dec_proto=OPRF+Take；不单独记录 Ext；"
        "通信量=论文字段（OPRF:id,u,v；Give:id,t,ct,τ,e；Take:id,t,ct,τ；"
        "S3:id,pwd,rid,sid），不含帧头。"
    )
    phases_spec = [
        (PHASE_REG, "reg_ms", "reg_comm", 1),
        (PHASE_ENC, "enc_proto_ms", "enc_proto_comm", 2),
        (PHASE_DEC, "dec_proto_ms", "dec_proto_comm", 2),
    ]
    agg = {}
    for name, lat_k, comm_k, rounds in phases_spec:
        lm, ls = mean_std([r[lat_k] for r in rows])
        cm, cs = mean_std([r[comm_k] for r in rows])
        agg[name] = {
            "rounds": rounds,
            "lat_mean": lm,
            "lat_std": ls,
            "comm_mean": cm,
            "comm_std": cs,
        }

    md_lines = [
        "# E2SE TCP 基准（对标 PAEE Init/Rec）",
        "",
        f"生成时间: {ts}",
        f"trials={trials} success={len(rows)} ({100.0 * len(rows) / max(trials, 1):.1f}%)",
        f"Server: `{endpoint}`",
        f"S3: {s3_mode}",
        "",
        "Enc_proto=OPRF+Give；Dec_proto=OPRF+Take；Reg 一次性参考。",
        "",
        "| 阶段 | 轮数 | 延迟 mean±std (ms) | 通信量 mean±std (B) |",
        "|------|------|--------------------|---------------------|",
    ]
    for name, _, _, _ in phases_spec:
        a = agg[name]
        md_lines.append(
            f"| {name} | {a['rounds']} | "
            f"{a['lat_mean']:.2f} ± {a['lat_std']:.2f} | "
            f"{a['comm_mean']:.1f} ± {a['comm_std']:.1f} |"
        )
    md_lines.extend(
        [
            "",
            "## 阶段说明",
            "",
            "| 阶段 | E2SE |",
            "|------|------|",
            "| Reg | Register |",
            "| Enc_proto≈Init | OPRF+Give |",
            "| Dec_proto≈Rec | OPRF+Take |",
            "",
            "## 说明",
            "",
            f"- {note}",
            "- **通信量**含 Key Server 字段 + S3 字段（id/pwd/rid/sid）；延迟在 `--real-s3` 时另含 S3 RTT。",
            "- 不单独记录 PAEE Ext；OPRF 已计入 Enc_proto / Dec_proto。",
            "",
        ]
    )
    md_path = out_dir / "tcp_network_benchmark.md"
    md_path.write_text("\n".join(md_lines), encoding="utf-8")
    (out_dir / "key_proto_comparison.md").write_text("\n".join(md_lines), encoding="utf-8")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        print("[bench] openpyxl missing; skip xlsx")
        return md_path, None

    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"
    headers = [
        "方案", "阶段", "试验次数", "成功次数", "成功率(%)", "轮数",
        "延迟_mean(ms)", "延迟_std(ms)", "通信量_mean(bytes)", "通信量_std(bytes)",
        "Server 端点", "测量说明", "时间戳",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    label_map = {
        PHASE_ENC: "封装≈Init(OPRF+Give)",
        PHASE_DEC: "恢复≈Rec(OPRF+Take)",
        PHASE_REG: "Reg(一次性)",
    }
    for name, _, _, rounds in phases_spec:
        a = agg[name]
        for label in dict.fromkeys([label_map[name], name]):
            ws.append([
                DISPLAY, label, trials, len(rows),
                round(100.0 * len(rows) / max(trials, 1), 1), rounds,
                round(a["lat_mean"], 3), round(a["lat_std"], 3),
                round(a["comm_mean"], 1), round(a["comm_std"], 1),
                endpoint, note, ts,
            ])

    ws2 = wb.create_sheet("原始数据")
    ws2.append([
        "方案", "trial_id", "reg_ms", "reg_B",
        "enc_proto_ms", "enc_proto_B", "dec_proto_ms", "dec_proto_B",
    ])
    for i, r in enumerate(rows, start=1):
        ws2.append([
            DISPLAY, i,
            round(r["reg_ms"], 3), round(r["reg_comm"], 1),
            round(r["enc_proto_ms"], 3), round(r["enc_proto_comm"], 1),
            round(r["dec_proto_ms"], 3), round(r["dec_proto_comm"], 1),
        ])

    xlsx_path = out_dir / "tcp_network_benchmark.xlsx"
    wb.save(xlsx_path)
    wb.save(out_dir / "key_proto_benchmark.xlsx")
    return md_path, xlsx_path


def main() -> int:
    ap = argparse.ArgumentParser(description="E2SE 分阶段 TCP 基准（对齐 PAEE）")
    ap.add_argument("--trials", type=int, default=20)
    ap.add_argument("--warmup", type=int, default=1)
    ap.add_argument("--auto-server", action="store_true")
    ap.add_argument("--host", default=None)
    ap.add_argument("--port", type=int, default=None)
    ap.add_argument(
        "--real-s3",
        action="store_true",
        help="使用 config.properties 的真实 AWS S3（延迟含 S3 RTT）",
    )
    args = ap.parse_args()

    try:
        import Constants
        host = args.host or Constants.AUTH_SERVER_ADDRESS or DEFAULT_HOST
        port = args.port or Constants.AUTH_SERVER_PORT_NUMBER or DEFAULT_PORT
    except Exception:
        host = args.host or DEFAULT_HOST
        port = args.port or DEFAULT_PORT

    if args.host is None and args.auto_server:
        host = DEFAULT_HOST

    proc: Optional[subprocess.Popen] = None
    try:
        if args.auto_server:
            print(f"[bench] AuthServer {host}:{port} ...")
            proc = start_auth_server(host, port)
        elif not wait_port(host, port, timeout_s=2.0):
            print(f"[bench] no server on {host}:{port}; use --auto-server 或先启动 AuthServer")
            return 1

        make_client(real_s3=args.real_s3, quiet_mode_print=True)
        print(
            f"[bench] trials={args.trials} warmup={args.warmup} "
            f"server={host}:{port} real_s3={args.real_s3}"
        )

        rows: List[Dict[str, float]] = []
        total = args.warmup + args.trials
        for i in range(1, total + 1):
            client = make_client(real_s3=args.real_s3, quiet_mode_print=False)
            try:
                row = run_trial(client, i, host, port)
            except Exception as e:
                print(f"  [fail {i}] {e}")
                continue
            tag = "warmup" if i <= args.warmup else f"{i - args.warmup}/{args.trials}"
            print(
                f"  [{tag}] "
                f"reg={row['reg_ms']:.2f}ms/{row['reg_comm']:.0f}B  "
                f"enc={row['enc_proto_ms']:.2f}ms/{row['enc_proto_comm']:.0f}B  "
                f"dec={row['dec_proto_ms']:.2f}ms/{row['dec_proto_comm']:.0f}B"
            )
            if i > args.warmup:
                rows.append(row)

        if not rows:
            print("[bench] no successful trials")
            return 1

        def show(name: str, lat_k: str, comm_k: str) -> None:
            lm, ls = mean_std([r[lat_k] for r in rows])
            cm, cs = mean_std([r[comm_k] for r in rows])
            print(f"  {name:18s}  {lm:8.2f} ± {ls:6.2f} ms  |  {cm:7.1f} ± {cs:5.1f} B")

        print()
        print("========== E2SE 分阶段结果（对齐 PAEE） ==========")
        if args.real_s3:
            print("（延迟含真实 S3；通信量仍仅为 AuthServer）")
        show(PHASE_REG, "reg_ms", "reg_comm")
        show(PHASE_ENC, "enc_proto_ms", "enc_proto_comm")
        show(PHASE_DEC, "dec_proto_ms", "dec_proto_comm")

        md, xlsx = write_outputs(
            ROOT / "experiment" / "output",
            rows,
            args.trials,
            host,
            port,
            real_s3=args.real_s3,
        )
        print(f"\n报告: {md}")
        if xlsx:
            print(f"Excel: {xlsx}")
        return 0
    finally:
        if proc is not None:
            proc.terminate()
            try:
                proc.wait(timeout=3)
            except subprocess.TimeoutExpired:
                proc.kill()
            logf = getattr(proc, "_bench_logf", None)
            if logf is not None:
                logf.close()


if __name__ == "__main__":
    raise SystemExit(main())
