"""实验三：口令长度固定 16，不同文件大小下的 Enc/Dec 延迟（对齐 PAEE）。

对齐 PassCrypt-PAEE ``file_size_proto_bench``::

    PAEE Enc_total = Ext + 封装存钥 + 本地加密文件
    PPKR Enc_total = Init（TCP 封装存钥）+ 本地 AES 加密文件
    PAEE Dec_total = Ext+Dec→dek + 本地解密
    PPKR Dec_total = Rec（TCP 恢复密钥）+ 本地 AES 解密

大文件密文**不上传**（仅本机）。口令长度默认 16。

文件大小默认：1, 10, 100, 200, 300, 400, 500 MB。
输出表：行=阶段，列=文件大小。

用法::

    python -m experiment.file_size_proto_bench --host <ServerIP> --port 8765 --trials 3
    python -m experiment.file_size_proto_bench --trials 3 --auto-server --sizes 1 10
"""

from __future__ import annotations

import argparse
import csv
import secrets
import statistics
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Sequence

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from client.ppkr_http_client import EncPwPlusHttpSession, OPRFPPKRHttpSession
from common.endpoint import DEFAULT_HOST, DEFAULT_PORT, check_server as tcp_check, resolve_endpoint
from crypto.aes_gcm import AESGCMCipher
from experiment.http_benchmark import start_auto_server
from logging_config import setup_logger

log = setup_logger("FILE-SIZE")

DEFAULT_SIZES_MB = (1, 10, 100, 200, 300, 400, 500)
PASSWORD_LEN = 16

PROTOCOL_META = {
    "encpw_plus": "π_encPw+ (Lev-2)",
    "oprf_ppkr": "π_OPRF-PPKR (Lev-3)",
}

# 主表 / 分解（对齐 PAEE 行结构；PPKR 无 Ext）
PHASE_ROWS = (
    ("Init≈Enc_proto（TCP 封装存钥）", "init_ms"),
    ("Enc（Init+本地加密文件）", "enc_total_ms"),
    ("Dec（Rec+本地解密文件）", "dec_total_ms"),
)

BREAKDOWN_ROWS = (
    ("Init（TCP）", "init_ms"),
    ("Enc_file（本地 AES.Enc）", "enc_file_ms"),
    ("Enc_total（Init+file）", "enc_total_ms"),
    ("Rec（TCP）", "rec_ms"),
    ("Dec_file（本地 AES.Dec）", "dec_file_ms"),
    ("Dec_total（Rec+file）", "dec_total_ms"),
)


def make_password(length: int = PASSWORD_LEN) -> str:
    return "a" * length


def _mean_std(values: list[float]) -> tuple[float, float]:
    if not values:
        return 0.0, 0.0
    if len(values) == 1:
        return values[0], 0.0
    return statistics.mean(values), statistics.stdev(values)


def _gen(size_bytes: int) -> bytes:
    import os

    chunk = 8 * 1024 * 1024
    parts: list[bytes] = []
    left = size_bytes
    while left > 0:
        n = min(chunk, left)
        parts.append(os.urandom(n))
        left -= n
    return b"".join(parts)


def _make_session(protocol: str, idc: str, password: str, host: str, port: int):
    if protocol == "encpw_plus":
        return EncPwPlusHttpSession(idc=idc, password=password, host=host, port=port)
    return OPRFPPKRHttpSession(idc=idc, password=password, host=host, port=port)


def run_size_trials(
    *,
    protocol: str,
    host: str,
    port: int,
    password: str,
    size_mb: int,
    trials: int,
    id_prefix: str,
    save_ct_dir: Path | None,
) -> list[dict[str, Any]]:
    """
    每个 trial（对齐 PAEE）::
      Init(TCP) → 本地 SE.Enc(file) → Rec(TCP) → 本地 SE.Dec
    密文不上 Server。
    """
    cipher = AESGCMCipher()
    size = size_mb * 1024 * 1024
    log.info("生成明文 %d MB ...", size_mb)
    plaintext = _gen(size)
    rows: list[dict[str, Any]] = []
    display = PROTOCOL_META[protocol]

    for t in range(1, trials + 1):
        idc = f"{id_prefix}_{size_mb}mb_{t}"
        log.info("---- %s %d MB trial %d/%d idc=%s ----", protocol, size_mb, t, trials, idc)
        session = None
        try:
            session = _make_session(protocol, idc, password, host, port)
            K, init_m = session.run_init_metrics()
            init_ms = float(init_m["latency_ms"])
            init_comm = int(init_m["comm_bytes"])

            t0 = time.perf_counter()
            file_ct, _ = cipher.enc(K, plaintext)
            enc_file_ms = (time.perf_counter() - t0) * 1000.0

            if save_ct_dir is not None:
                save_ct_dir.mkdir(parents=True, exist_ok=True)
                (save_ct_dir / f"{idc}.ct").write_bytes(file_ct.serialize())

            enc_total_ms = init_ms + enc_file_ms

            K_rec, rec_m = session.run_rec_metrics()
            if K_rec != K:
                raise RuntimeError("Init/Rec 密钥不一致")
            rec_ms = float(rec_m["latency_ms"])
            rec_comm = int(rec_m["comm_bytes"])

            t1 = time.perf_counter()
            out, _ = cipher.dec(K_rec, file_ct)
            dec_file_ms = (time.perf_counter() - t1) * 1000.0
            if out != plaintext:
                raise RuntimeError("plaintext mismatch")

            dec_total_ms = rec_ms + dec_file_ms

            row = {
                "success": True,
                "protocol": protocol,
                "display": display,
                "size_mb": size_mb,
                "trial_id": t,
                "idc": idc,
                "init_ms": init_ms,
                "enc_file_ms": enc_file_ms,
                "enc_total_ms": enc_total_ms,
                "rec_ms": rec_ms,
                "dec_file_ms": dec_file_ms,
                "dec_total_ms": dec_total_ms,
                "init_comm_bytes": init_comm,
                "rec_comm_bytes": rec_comm,
                "ct_bytes": len(file_ct.serialize()),
                "error": "",
            }
            rows.append(row)
            log.info(
                "  Init=%.1f Enc=%.1f(init+file=%.1f) "
                "Dec=%.1f(rec=%.1f+file=%.1f) init_comm=%dB rec_comm=%dB",
                init_ms,
                enc_total_ms,
                enc_file_ms,
                dec_total_ms,
                rec_ms,
                dec_file_ms,
                init_comm,
                rec_comm,
            )
            del file_ct, out
        except Exception as e:  # noqa: BLE001
            log.error("trial fail: %s", e)
            rows.append(
                {
                    "success": False,
                    "protocol": protocol,
                    "display": display,
                    "size_mb": size_mb,
                    "trial_id": t,
                    "idc": idc,
                    "init_ms": 0.0,
                    "enc_file_ms": 0.0,
                    "enc_total_ms": 0.0,
                    "rec_ms": 0.0,
                    "dec_file_ms": 0.0,
                    "dec_total_ms": 0.0,
                    "init_comm_bytes": 0,
                    "rec_comm_bytes": 0,
                    "ct_bytes": 0,
                    "error": str(e),
                }
            )
        finally:
            if session is not None:
                session.close()

    del plaintext
    return rows


def summarize_by_size(
    all_raw: list[dict[str, Any]], sizes: Sequence[int]
) -> dict[int, dict[str, Any]]:
    out: dict[int, dict[str, Any]] = {}
    for mb in sizes:
        rows = [r for r in all_raw if r["size_mb"] == mb]
        ok = [r for r in rows if r["success"]]
        stats: dict[str, Any] = {
            "n_trials": len(rows),
            "n_success": len(ok),
        }
        for _name, key in BREAKDOWN_ROWS:
            m, s = _mean_std([float(r[key]) for r in ok])
            stats[f"{key}_mean"] = m
            stats[f"{key}_std"] = s
        for key in ("init_comm_bytes", "rec_comm_bytes"):
            m, s = _mean_std([float(r[key]) for r in ok])
            stats[f"{key}_mean"] = m
            stats[f"{key}_std"] = s
        out[mb] = stats
    return out


def _size_header(sizes: Sequence[int]) -> list[str]:
    return ["阶段"] + [f"{mb}MB" for mb in sizes]


def _row_mean_std(
    by_size: dict[int, dict[str, Any]],
    sizes: Sequence[int],
    phase: str,
    key: str,
) -> list[Any]:
    cells: list[Any] = [phase]
    for mb in sizes:
        st = by_size[mb]
        cells.append(f"{st[f'{key}_mean']:.2f} ± {st[f'{key}_std']:.2f}")
    return cells


def write_outputs(
    by_size: dict[int, dict[str, Any]],
    raw: list[dict[str, Any]],
    sizes: Sequence[int],
    out_dir: Path,
    *,
    display: str,
    endpoint: str,
    trials: int,
    password_len: int,
    timestamp: str,
) -> dict[str, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    note = (
        f"方案={display}；口令长度={password_len}；trials/大小={trials}；"
        f"Server={endpoint}；"
        "Enc_total=Init(TCP)+本地加密文件（大文件密文不上传）；"
        "Dec_total=Rec(TCP)+本地解密；列=文件大小。对齐 PAEE 实验三。"
    )
    header = _size_header(sizes)
    sep = "|" + "|".join(["------"] * len(header)) + "|"

    md = [
        "# PPKR 实验三：口令长度=16，不同文件大小下 Enc/Dec 延迟",
        "",
        f"生成时间: {timestamp}",
        note,
        "",
        "## 主表 mean±std (ms)",
        "",
        "| " + " | ".join(header) + " |",
        sep,
    ]
    for phase, key in PHASE_ROWS:
        md.append(
            "| "
            + " | ".join(str(c) for c in _row_mean_std(by_size, sizes, phase, key))
            + " |"
        )
    md += ["", "## 分解 mean±std (ms)", "", "| " + " | ".join(header) + " |", sep]
    for phase, key in BREAKDOWN_ROWS:
        md.append(
            "| "
            + " | ".join(str(c) for c in _row_mean_std(by_size, sizes, phase, key))
            + " |"
        )
    md += [
        "",
        "## 协议通信量 mean±std (B)（与文件大小无关，大文件密文不上网）",
        "",
        "| " + " | ".join(header) + " |",
        sep,
    ]
    for phase, key in (
        ("Init 通信量", "init_comm_bytes"),
        ("Rec 通信量", "rec_comm_bytes"),
    ):
        md.append(
            "| "
            + " | ".join(str(c) for c in _row_mean_std(by_size, sizes, phase, key))
            + " |"
        )

    md_path = out_dir / "file_size_proto_benchmark.md"
    md_path.write_text("\n".join(md) + "\n", encoding="utf-8")

    summary_csv = out_dir / "file_size_proto_benchmark_summary.csv"
    with summary_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        for phase, key in PHASE_ROWS:
            w.writerow(_row_mean_std(by_size, sizes, phase, key))
        w.writerow([])
        w.writerow(["分解"])
        for phase, key in BREAKDOWN_ROWS:
            w.writerow(_row_mean_std(by_size, sizes, phase, key))
        w.writerow([])
        w.writerow(["通信量(B)"])
        for phase, key in (
            ("Init 通信量", "init_comm_bytes"),
            ("Rec 通信量", "rec_comm_bytes"),
        ):
            w.writerow(_row_mean_std(by_size, sizes, phase, key))

    raw_csv = out_dir / "file_size_proto_benchmark.csv"
    if raw:
        with raw_csv.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(raw[0].keys()))
            w.writeheader()
            w.writerows(raw)

    def style_h(ws) -> None:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"
    ws.append(header)
    style_h(ws)
    for phase, key in PHASE_ROWS:
        ws.append(_row_mean_std(by_size, sizes, phase, key))
    ws.append([])
    ws.append([note])
    ws.append([f"方案={display}", f"时间={timestamp}"])

    ws2 = wb.create_sheet("分解")
    ws2.append(header)
    style_h(ws2)
    for phase, key in BREAKDOWN_ROWS:
        ws2.append(_row_mean_std(by_size, sizes, phase, key))

    ws3 = wb.create_sheet("延迟_mean")
    ws3.append(header)
    style_h(ws3)
    for phase, key in PHASE_ROWS:
        ws3.append([phase] + [round(by_size[mb][f"{key}_mean"], 3) for mb in sizes])

    ws_comm = wb.create_sheet("通信量")
    ws_comm.append(header)
    style_h(ws_comm)
    for phase, key in (
        ("Init 通信量", "init_comm_bytes"),
        ("Rec 通信量", "rec_comm_bytes"),
    ):
        ws_comm.append(_row_mean_std(by_size, sizes, phase, key))

    ws4 = wb.create_sheet("原始数据")
    if raw:
        keys = list(raw[0].keys())
        ws4.append(keys)
        style_h(ws4)
        for r in raw:
            ws4.append([r[k] for k in keys])

    xlsx_path = out_dir / "file_size_proto_benchmark.xlsx"
    wb.save(xlsx_path)
    log.info("wrote %s", xlsx_path)
    return {"md": md_path, "csv": raw_csv, "summary_csv": summary_csv, "xlsx": xlsx_path}


def main() -> int:
    parser = argparse.ArgumentParser(
        description="PPKR 实验三：口令长度=16，不同文件大小 Enc/Dec 延迟（对齐 PAEE）"
    )
    parser.add_argument("--trials", type=int, default=3)
    parser.add_argument(
        "--sizes",
        type=int,
        nargs="+",
        default=list(DEFAULT_SIZES_MB),
        help="文件大小 MB，默认 1 10 100 200 300 400 500",
    )
    parser.add_argument("--password-len", type=int, default=PASSWORD_LEN)
    parser.add_argument("--host", default=DEFAULT_HOST)
    parser.add_argument("--port", type=int, default=DEFAULT_PORT)
    parser.add_argument("--url", default=None)
    parser.add_argument("--auto-server", action="store_true")
    parser.add_argument(
        "--protocol",
        default="oprf_ppkr",
        choices=list(PROTOCOL_META.keys()),
        help="默认 π_OPRF-PPKR",
    )
    parser.add_argument(
        "--save-ct",
        action="store_true",
        help="将文件密文写到 out-dir/file_cts/（不计时）",
    )
    parser.add_argument(
        "--out-dir", default=str(ROOT / "experiment" / "output")
    )
    args = parser.parse_args()

    password = make_password(args.password_len)
    assert len(password) == args.password_len

    server = None
    if args.auto_server:
        host, port, server = start_auto_server()
    else:
        host, port = resolve_endpoint(host=args.host, port=args.port, url=args.url)
        tcp_check(host, port)

    run_tag = datetime.now().strftime("%Y%m%d%H%M%S") + secrets.token_hex(2)
    id_prefix = f"fsize_{args.protocol}_{run_tag}"
    save_ct_dir = (
        Path(args.out_dir) / "file_cts" / run_tag if args.save_ct else None
    )
    display = PROTOCOL_META[args.protocol]

    all_raw: list[dict[str, Any]] = []
    try:
        for mb in args.sizes:
            log.info("===== file size %d MB =====", mb)
            all_raw.extend(
                run_size_trials(
                    protocol=args.protocol,
                    host=host,
                    port=port,
                    password=password,
                    size_mb=mb,
                    trials=args.trials,
                    id_prefix=id_prefix,
                    save_ct_dir=save_ct_dir,
                )
            )

        by_size = summarize_by_size(all_raw, args.sizes)
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        paths = write_outputs(
            by_size,
            all_raw,
            args.sizes,
            Path(args.out_dir),
            display=display,
            endpoint=f"{host}:{port}",
            trials=args.trials,
            password_len=args.password_len,
            timestamp=ts,
        )

        header = _size_header(args.sizes)
        col_w = max(16, max(len(h) for h in header[1:]) + 2)
        print()
        print(f"=== 实验三 {display}（口令长度={args.password_len}）===")
        print(f"{header[0]:<42}" + "".join(f"{h:>{col_w}}" for h in header[1:]))
        for phase, key in PHASE_ROWS:
            row = _row_mean_std(by_size, args.sizes, phase, key)
            print(f"{row[0]:<42}" + "".join(f"{c:>{col_w}}" for c in row[1:]))
        print(f"\nMarkdown: {paths['md']}")
        print(f"Excel:    {paths['xlsx']}")
    finally:
        if server is not None:
            server.stop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
