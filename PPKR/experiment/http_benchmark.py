"""实验一：测量 PPKR 各阶段的运行时间与通信量（TCP）。

对齐 PassCrypt-PAEE 实验一（``tcp_benchmark``）口径与报告结构::

    PAEE Enc_proto ≈ PPKR Init   （封装 DEK）
    PAEE Dec_proto ≈ PPKR Rec    （恢复 DEK）

测量内容（Init / Rec 分别统计，默认 20 次取平均）::
    - 运行时间（ms）：含 Client↔Server TCP 往返 + 本地密码学
    - 通信量（bytes）：该阶段全部 TCP wire（含 4 字节长度头；不含 HELLO）
    - 轮数：Fig.1 交互轮（一对 C↔S 算 1；不含 HELLO / 本地 finish）

用法::

    # 终端1
    python serve.py --port 8765

    # 终端2
    python -m experiment.tcp_benchmark --trials 20 --port 8765

    # 或一键拉起临时 Server
    python -m experiment.tcp_benchmark --trials 20 --auto-server

输出（``experiment/output/``）::
    tcp_network_benchmark.md / .csv / .xlsx
"""

from __future__ import annotations

import argparse
import csv
import socket
import statistics
import sys
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from client.ppkr_http_client import EncPwPlusHttpSession, OPRFPPKRHttpSession
from common.endpoint import DEFAULT_HOST, DEFAULT_PORT, check_server as tcp_check, resolve_endpoint
from logging_config import setup_logger

log = setup_logger("TCP-BENCH")

# 一轮 = 一次 Client↔Server 交互（Fig.1 一对箭头）
PROTOCOL_META = {
    "encpw_plus": {"display": "π_encPw+ (Lev-2)", "rounds_init": 2, "rounds_rec": 2},
    "oprf_ppkr": {"display": "π_OPRF-PPKR (Lev-3)", "rounds_init": 2, "rounds_rec": 2},
}

NOTE = (
    "口径对齐 PAEE 实验一：Init≈Enc_proto（封装 DEK），Rec≈Dec_proto（恢复 DEK）；"
    "延迟含 TCP 往返+本地密码学；通信量=wire（binary opcode 帧，不含 HELLO）；"
    "轮数=Fig.1 交互轮（一对 C↔S 算 1；不含 HELLO/本地 finish）。"
)


def _make_session(protocol: str, idc: str, password: str, host: str, port: int):
    """按协议名创建 TCP 会话（EncPwPlus 或 OPRF-PPKR）。"""
    if protocol == "encpw_plus":
        return EncPwPlusHttpSession(idc=idc, password=password, host=host, port=port)
    return OPRFPPKRHttpSession(idc=idc, password=password, host=host, port=port)


def _mean_std(values: list[float]) -> tuple[float, float]:
    if not values:
        return 0.0, 0.0
    if len(values) == 1:
        return values[0], 0.0
    return statistics.mean(values), statistics.stdev(values)


def check_server(host: str, port: int) -> None:
    """探测 TCP Server 是否可达。"""
    tcp_check(host, port)


def _free_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind(("127.0.0.1", 0))
        return int(s.getsockname()[1])


def start_auto_server() -> tuple[str, int, Any]:
    """后台启动临时 TcpServer，返回 (host, port, server)。"""
    from server.tcp_server import TcpServer

    host = "127.0.0.1"
    port = _free_port()
    server = TcpServer(host, port)
    threading.Thread(target=server.start, daemon=True).start()
    deadline = time.time() + 8
    while time.time() < deadline:
        try:
            with socket.create_connection((host, port), timeout=0.2):
                log.info("auto-server ready %s:%d", host, port)
                return host, port, server
        except OSError:
            time.sleep(0.05)
    raise SystemExit("auto-server 启动失败")


def run_protocol_trials(
    protocol: str, n: int, password: str, host: str, port: int, idc_prefix: str
) -> list[dict]:
    """对单个协议跑 n 次完整 Init+Rec（TCP），返回原始 trial 列表。"""
    rows: list[dict] = []
    meta = PROTOCOL_META[protocol]
    for i in range(1, n + 1):
        idc = f"{idc_prefix}_{i}"
        log.info("---------- %s trial %d/%d idc=%s ----------", protocol, i, n, idc)
        session = None
        try:
            session = _make_session(protocol, idc, password, host, port)
            K_init, init_m = session.run_init_metrics()
            K_rec, rec_m = session.run_rec_metrics()
            key_match = K_init == K_rec
            rows.append(
                {
                    "protocol": protocol,
                    "display": meta["display"],
                    "trial_id": i,
                    "idc": idc,
                    "success": key_match,
                    "key_match": key_match,
                    "init_latency_ms": init_m["latency_ms"],
                    "rec_latency_ms": rec_m["latency_ms"],
                    "total_latency_ms": init_m["latency_ms"] + rec_m["latency_ms"],
                    "init_comm_bytes": init_m["comm_bytes"],
                    "rec_comm_bytes": rec_m["comm_bytes"],
                    "total_comm_bytes": init_m["comm_bytes"] + rec_m["comm_bytes"],
                    "init_rounds": meta["rounds_init"],
                    "rec_rounds": meta["rounds_rec"],
                    "error": "",
                }
            )
            log.info(
                "trial %d OK init=%.2fms/%dB rec=%.2fms/%dB match=%s",
                i,
                init_m["latency_ms"],
                init_m["comm_bytes"],
                rec_m["latency_ms"],
                rec_m["comm_bytes"],
                key_match,
            )
        except Exception as e:
            log.error("trial %d 失败: %s", i, e)
            rows.append(
                {
                    "protocol": protocol,
                    "display": meta["display"],
                    "trial_id": i,
                    "idc": idc,
                    "success": False,
                    "key_match": False,
                    "init_latency_ms": 0.0,
                    "rec_latency_ms": 0.0,
                    "total_latency_ms": 0.0,
                    "init_comm_bytes": 0,
                    "rec_comm_bytes": 0,
                    "total_comm_bytes": 0,
                    "init_rounds": meta["rounds_init"],
                    "rec_rounds": meta["rounds_rec"],
                    "error": str(e),
                }
            )
        finally:
            if session is not None:
                session.close()
    return rows


def aggregate(protocol: str, rows: list[dict]) -> dict:
    """对成功 trial 取平均；phases 结构对齐 PAEE 实验一。"""
    meta = PROTOCOL_META[protocol]
    ok = [r for r in rows if r["success"]]

    def pack(lat_key: str, comm_key: str) -> dict:
        lm, ls = _mean_std([float(r[lat_key]) for r in ok])
        cm, cs = _mean_std([float(r[comm_key]) for r in ok])
        return {
            "latency_mean_ms": lm,
            "latency_std_ms": ls,
            "comm_mean_bytes": cm,
            "comm_std_bytes": cs,
        }

    init_p = pack("init_latency_ms", "init_comm_bytes")
    rec_p = pack("rec_latency_ms", "rec_comm_bytes")
    tot_p = pack("total_latency_ms", "total_comm_bytes")

    phases = {
        "Init≈Enc_proto": {**init_p, "rounds": meta["rounds_init"]},
        "Rec≈Dec_proto": {**rec_p, "rounds": meta["rounds_rec"]},
        "Total": {
            **tot_p,
            "rounds": meta["rounds_init"] + meta["rounds_rec"],
        },
    }

    return {
        "protocol": protocol,
        "display": meta["display"],
        "n_trials": len(rows),
        "n_success": len(ok),
        "success_rate": len(ok) / len(rows) if rows else 0.0,
        "init_rounds": meta["rounds_init"],
        "rec_rounds": meta["rounds_rec"],
        # 兼容旧 Excel / compare 脚本字段名
        "init_latency_mean_ms": init_p["latency_mean_ms"],
        "init_latency_std_ms": init_p["latency_std_ms"],
        "rec_latency_mean_ms": rec_p["latency_mean_ms"],
        "rec_latency_std_ms": rec_p["latency_std_ms"],
        "total_latency_mean_ms": tot_p["latency_mean_ms"],
        "total_latency_std_ms": tot_p["latency_std_ms"],
        "init_comm_mean_bytes": init_p["comm_mean_bytes"],
        "init_comm_std_bytes": init_p["comm_std_bytes"],
        "rec_comm_mean_bytes": rec_p["comm_mean_bytes"],
        "rec_comm_std_bytes": rec_p["comm_std_bytes"],
        "total_comm_mean_bytes": tot_p["comm_mean_bytes"],
        "total_comm_std_bytes": tot_p["comm_std_bytes"],
        "phases": phases,
    }


def write_excel(
    aggregates: list[dict],
    raw_rows: list[dict],
    output_path: Path,
    meta: dict,
) -> Path:
    """写入 Excel：汇总 + 原始数据（列布局兼容 PAEE/跨方案对比脚本）。"""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "汇总"
    headers = [
        "方案",
        "阶段",
        "试验次数",
        "成功次数",
        "成功率(%)",
        "轮数",
        "延迟_mean(ms)",
        "延迟_std(ms)",
        "通信量_mean(bytes)",
        "通信量_std(bytes)",
        "Server 端点",
        "测量说明",
        "时间戳",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Excel 阶段名保持 Init/Rec/Total，便于 compare_paee_ppkr_wbp 读取
    excel_phase_map = {
        "Init≈Enc_proto": "Init",
        "Rec≈Dec_proto": "Rec",
        "Total": "Total",
    }
    for agg in aggregates:
        for phase_key, p in agg["phases"].items():
            ws.append(
                [
                    agg["display"],
                    excel_phase_map.get(phase_key, phase_key),
                    agg["n_trials"],
                    agg["n_success"],
                    round(agg["success_rate"] * 100, 1),
                    p["rounds"],
                    round(p["latency_mean_ms"], 3),
                    round(p["latency_std_ms"], 3),
                    round(p["comm_mean_bytes"], 1),
                    round(p["comm_std_bytes"], 1),
                    meta["url"],
                    NOTE,
                    meta["timestamp"],
                ]
            )

    ws2 = wb.create_sheet("原始数据")
    raw_headers = [
        "方案",
        "trial_id",
        "idc",
        "success",
        "key_match",
        "Init延迟(ms)",
        "Rec延迟(ms)",
        "Total延迟(ms)",
        "Init通信量(bytes)",
        "Rec通信量(bytes)",
        "Total通信量(bytes)",
        "Init轮数",
        "Rec轮数",
        "error",
    ]
    ws2.append(raw_headers)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for r in raw_rows:
        ws2.append(
            [
                r["display"],
                r["trial_id"],
                r["idc"],
                r["success"],
                r["key_match"],
                round(r["init_latency_ms"], 3),
                round(r["rec_latency_ms"], 3),
                round(r["total_latency_ms"], 3),
                r["init_comm_bytes"],
                r["rec_comm_bytes"],
                r["total_comm_bytes"],
                r["init_rounds"],
                r["rec_rounds"],
                r["error"],
            ]
        )

    for sheet in (ws, ws2):
        for col in sheet.columns:
            letter = col[0].column_letter
            width = min(40, max(12, max(len(str(c.value or "")) for c in col) + 2))
            sheet.column_dimensions[letter].width = width

    wb.save(output_path)
    return output_path


def write_markdown(
    aggregates: list[dict],
    out_path: Path,
    *,
    endpoint: str,
    timestamp: str,
) -> Path:
    """写入 PAEE 风格 Markdown 阶段表。"""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# PPKR 实验一：各阶段运行时间与通信量",
        "",
        f"生成时间: {timestamp}",
        f"Server: `{endpoint}`",
        "",
        NOTE,
        "",
    ]
    for agg in aggregates:
        lines.extend(
            [
                f"## {agg['display']}",
                "",
                f"trials={agg['n_trials']} success={agg['n_success']} "
                f"({agg['success_rate'] * 100:.1f}%)",
                "",
                "| 阶段 | 轮数 | 延迟 mean±std (ms) | 通信量 mean±std (B) |",
                "|------|------|--------------------|---------------------|",
            ]
        )
        for phase, p in agg["phases"].items():
            lines.append(
                f"| {phase} | {p['rounds']} | "
                f"{p['latency_mean_ms']:.2f} ± {p['latency_std_ms']:.2f} | "
                f"{p['comm_mean_bytes']:.1f} ± {p['comm_std_bytes']:.1f} |"
            )
        lines.append("")
    out_path.write_text("\n".join(lines), encoding="utf-8")
    return out_path


def write_csv(raw_rows: list[dict], out_path: Path) -> Path:
    """写入原始 trial CSV。"""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if not raw_rows:
        out_path.write_text("", encoding="utf-8")
        return out_path
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(raw_rows[0].keys()))
        w.writeheader()
        w.writerows(raw_rows)
    return out_path


def print_summary(aggregates: list[dict], paths: dict[str, Path]) -> None:
    print("\n" + "=" * 72)
    print("  实验一：PPKR 各阶段运行时间与通信量（对齐 PAEE）")
    print("=" * 72)
    for agg in aggregates:
        print(f"\n【{agg['display']}】 n={agg['n_trials']} 成功={agg['n_success']}")
        for phase, p in agg["phases"].items():
            print(
                f"  {phase:16s}  "
                f"{p['latency_mean_ms']:8.2f} ± {p['latency_std_ms']:6.2f} ms  "
                f"{p['comm_mean_bytes']:8.1f} ± {p['comm_std_bytes']:6.1f} B  "
                f"(rounds={p['rounds']})"
            )
    print(f"\nMarkdown: {paths['md']}")
    print(f"CSV:      {paths['csv']}")
    print(f"Excel:    {paths['xlsx']}")
    print("=" * 72 + "\n")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="PPKR 实验一：各阶段运行时间与通信量 → MD/CSV/Excel"
    )
    parser.add_argument("--trials", type=int, default=20, help="每个协议试验次数（默认20）")
    parser.add_argument("--password", default="benchmark_pw_2024")
    parser.add_argument("--host", default=DEFAULT_HOST)
    parser.add_argument("--port", type=int, default=DEFAULT_PORT)
    parser.add_argument("--url", default=None, help="兼容写法 host:port")
    parser.add_argument(
        "--auto-server",
        action="store_true",
        help="自动在临时端口启动 TcpServer（无需另开终端）",
    )
    parser.add_argument(
        "--protocols",
        nargs="+",
        default=["oprf_ppkr"],
        help="默认仅测 π_OPRF-PPKR；对比两方案时加 encpw_plus",
        choices=list(PROTOCOL_META.keys()),
    )
    parser.add_argument(
        "--out-dir",
        type=str,
        default=str(ROOT / "experiment" / "output"),
        help="输出目录（默认 experiment/output）",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Excel 路径（默认 out-dir/tcp_network_benchmark.xlsx）",
    )
    args = parser.parse_args()

    server = None
    if args.auto_server:
        host, port, server = start_auto_server()
    else:
        host, port = resolve_endpoint(host=args.host, port=args.port, url=args.url)
        check_server(host, port)

    endpoint = f"{host}:{port}"
    log.info(
        "实验一 TCP 基准 trials=%d protocols=%s %s",
        args.trials,
        args.protocols,
        endpoint,
    )

    try:
        all_raw: list[dict] = []
        aggregates: list[dict] = []
        for name in args.protocols:
            rows = run_protocol_trials(
                name, args.trials, args.password, host, port, idc_prefix=f"tcp_{name}"
            )
            all_raw.extend(rows)
            aggregates.append(aggregate(name, rows))
    finally:
        if server is not None:
            server.stop()

    out_dir = Path(args.out_dir)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    xlsx_path = Path(args.output) if args.output else out_dir / "tcp_network_benchmark.xlsx"
    md_path = out_dir / "tcp_network_benchmark.md"
    csv_path = out_dir / "tcp_network_benchmark.csv"

    meta = {"timestamp": ts, "url": endpoint, "trials": args.trials}
    write_excel(aggregates, all_raw, xlsx_path, meta)
    write_markdown(aggregates, md_path, endpoint=endpoint, timestamp=ts)
    write_csv(all_raw, csv_path)
    print_summary(aggregates, {"md": md_path, "csv": csv_path, "xlsx": xlsx_path})


if __name__ == "__main__":
    main()
