"""实验设计 1（对齐 PAEE）：不同口令长度下 WBP Init / Rec 延迟。

口径与 ``tcp_benchmark`` 一致（Init/Rec 含 TCP；不含 HELLO；二进制 wire，无 JSON/Base64）。
口令为可打印单字符重复，仅长度变化。

默认长度：8, 16, 32, 64, 128, 256, 512。

用法::

    # 终端1（本机或 EC2）
    python run_server.py --host 0.0.0.0 --port 8765

    # 终端2
    python -m experiment.password_length_bench --trials 20 --host 127.0.0.1 --port 8765
    python -m experiment.password_length_bench --trials 20 --host 54.x.x.x --port 8765 -q

输出::
    experiment/output/password_length_benchmark.xlsx
    experiment/output/password_length_benchmark.md
    experiment/output/password_length_benchmark_summary.csv
    experiment/output/password_length_benchmark.csv
"""

from __future__ import annotations

import argparse
import csv
import logging
import secrets
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from client.app import DEFAULT_HOST, DEFAULT_PORT, INIT_ROUNDS, REC_ROUNDS
from experiment.tcp_benchmark import check_server, run_trials, _mean_std

log = logging.getLogger("wbp.pwlen")

DISPLAY = "WBP (DFG+23)"
DEFAULT_LENGTHS = (8, 16, 32, 64, 128, 256, 512)

# 行=阶段，列=口令长度（对齐 PAEE 实验1 表格式）
PHASES = (
    ("Init≈Enc_proto", "init_latency_ms", "init_comm_bytes", INIT_ROUNDS),
    ("Rec≈Dec_proto", "rec_latency_ms", "rec_comm_bytes", REC_ROUNDS),
)


def make_password(length: int, *, alphabet: str = "a") -> str:
    """构造指定字节长度的口令（UTF-8 单字节字符重复）。"""
    if length < 1:
        raise ValueError("password length must be >= 1")
    if len(alphabet.encode("utf-8")) != 1:
        raise ValueError("alphabet must be a single ASCII character")
    return alphabet * length


def parse_lengths(spec: str) -> list[int]:
    """解析 ``8,16,32`` 或 ``8:512``（2 的幂）或 ``8:512:32``（步长）。"""
    spec = spec.strip()
    if not spec:
        return list(DEFAULT_LENGTHS)
    if "," in spec:
        return [int(x) for x in spec.split(",") if x.strip()]
    if ":" in spec:
        parts = [int(x) for x in spec.split(":")]
        if len(parts) == 2:
            lo, hi = parts
            out: list[int] = []
            n = lo
            while n <= hi:
                out.append(n)
                n *= 2
            if out and hi not in out and hi >= lo:
                out.append(hi)
            return out
        if len(parts) == 3:
            lo, hi, step = parts
            return list(range(lo, hi + 1, step))
    return [int(spec)]


def summarize_length(length: int, rows: list[dict]) -> dict[str, Any]:
    ok = [r for r in rows if r["success"]]
    phases: dict[str, dict[str, float]] = {}
    for name, lat_key, comm_key, rounds in PHASES:
        lats = [float(r[lat_key]) for r in ok]
        comms = [float(r[comm_key]) for r in ok]
        lm, ls = _mean_std(lats)
        cm, cs = _mean_std(comms)
        phases[name] = {
            "rounds": rounds,
            "latency_mean_ms": lm,
            "latency_std_ms": ls,
            "comm_mean_bytes": cm,
            "comm_std_bytes": cs,
        }
    return {
        "password_length": length,
        "n_trials": len(rows),
        "n_success": len(ok),
        "success_rate": (len(ok) / len(rows)) if rows else 0.0,
        "phases": phases,
    }


def _length_header(summaries: list[dict[str, Any]]) -> list[Any]:
    return ["阶段"] + [f"L={s['password_length']}" for s in summaries]


def _phase_row_mean_std(
    summaries: list[dict[str, Any]], phase: str, *, metric: str = "latency"
) -> list[Any]:
    cells: list[Any] = [phase]
    for s in summaries:
        p = s["phases"][phase]
        if metric == "latency":
            cells.append(
                f"{p['latency_mean_ms']:.2f} ± {p['latency_std_ms']:.2f}"
            )
        else:
            cells.append(
                f"{p['comm_mean_bytes']:.1f} ± {p['comm_std_bytes']:.1f}"
            )
    return cells


def _phase_row_numeric(
    summaries: list[dict[str, Any]],
    phase: str,
    *,
    which: str,
    metric: str = "latency",
) -> list[Any]:
    if metric == "latency":
        key = "latency_mean_ms" if which == "mean" else "latency_std_ms"
    else:
        key = "comm_mean_bytes" if which == "mean" else "comm_std_bytes"
    cells: list[Any] = [phase]
    for s in summaries:
        cells.append(round(s["phases"][phase][key], 3))
    return cells


def write_outputs(
    summaries: list[dict[str, Any]],
    raw: list[dict[str, Any]],
    out_dir: Path,
    *,
    endpoint: str,
    trials: int,
    timestamp: str,
) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    note = (
        "实验1：口令长度扫描；口径同 tcp_benchmark；"
        "口令=单字符重复；Init/Rec 对标 PAEE Enc_proto/Dec_proto；"
        f"trials/长度={trials}；Server={endpoint}；"
        "表格式：行=阶段，列=口令长度"
    )
    header = _length_header(summaries)
    sep = "|" + "|".join(["------"] * len(header)) + "|"

    md_lines = [
        "# WBP 口令长度 vs Init / Rec 延迟",
        "",
        f"生成时间: {timestamp}",
        note,
        "",
        "## 延迟 mean±std (ms)",
        "",
        "| " + " | ".join(str(h) for h in header) + " |",
        sep,
    ]
    for name, *_rest in PHASES:
        md_lines.append(
            "| "
            + " | ".join(str(c) for c in _phase_row_mean_std(summaries, name))
            + " |"
        )
    md_lines += [
        "",
        "## 通信量 mean±std (bytes)",
        "",
        "| " + " | ".join(str(h) for h in header) + " |",
        sep,
    ]
    for name, *_rest in PHASES:
        md_lines.append(
            "| "
            + " | ".join(
                str(c)
                for c in _phase_row_mean_std(summaries, name, metric="comm")
            )
            + " |"
        )
    md_lines += [
        "",
        "## 成功次数 / 试验次数",
        "",
        "| " + " | ".join(str(h) for h in header) + " |",
        sep,
        "| "
        + " | ".join(
            ["成功/试验"]
            + [f"{s['n_success']}/{s['n_trials']}" for s in summaries]
        )
        + " |",
        "",
    ]
    (out_dir / "password_length_benchmark.md").write_text(
        "\n".join(md_lines), encoding="utf-8"
    )

    summary_csv = out_dir / "password_length_benchmark_summary.csv"
    with summary_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        for name, *_rest in PHASES:
            w.writerow(_phase_row_mean_std(summaries, name))
        w.writerow(
            ["成功/试验"]
            + [f"{s['n_success']}/{s['n_trials']}" for s in summaries]
        )

    if raw:
        csv_path = out_dir / "password_length_benchmark.csv"
        with csv_path.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(raw[0].keys()))
            w.writeheader()
            w.writerows(raw)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        log.warning("openpyxl missing; skip xlsx")
        return

    wb = Workbook()

    def _style_header(ws) -> None:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    ws = wb.active
    ws.title = "汇总"
    ws.append(header)
    _style_header(ws)
    for name, *_rest in PHASES:
        ws.append(_phase_row_mean_std(summaries, name))
    ws.append(
        ["成功/试验"]
        + [f"{s['n_success']}/{s['n_trials']}" for s in summaries]
    )
    ws.append([])
    ws.append([note])
    ws.append([f"方案={DISPLAY}", f"时间={timestamp}", f"端点={endpoint}"])

    ws_mean = wb.create_sheet("延迟_mean")
    ws_mean.append(header)
    _style_header(ws_mean)
    for name, *_rest in PHASES:
        ws_mean.append(_phase_row_numeric(summaries, name, which="mean"))

    ws_std = wb.create_sheet("延迟_std")
    ws_std.append(header)
    _style_header(ws_std)
    for name, *_rest in PHASES:
        ws_std.append(_phase_row_numeric(summaries, name, which="std"))

    ws_comm = wb.create_sheet("通信量_mean")
    ws_comm.append(header)
    _style_header(ws_comm)
    for name, *_rest in PHASES:
        ws_comm.append(
            _phase_row_numeric(summaries, name, which="mean", metric="comm")
        )

    ws3 = wb.create_sheet("原始数据")
    raw_headers = [
        "口令长度",
        "trial_id",
        "idc",
        "success",
        "Init延迟(ms)",
        "Rec延迟(ms)",
        "Init通信(B)",
        "Rec通信(B)",
        "error",
    ]
    ws3.append(raw_headers)
    _style_header(ws3)
    for r in raw:
        ws3.append(
            [
                r["password_length"],
                r["trial_id"],
                r["idc"],
                r["success"],
                round(r["init_latency_ms"], 3),
                round(r["rec_latency_ms"], 3),
                r["init_comm_bytes"],
                r["rec_comm_bytes"],
                r.get("error", ""),
            ]
        )

    path = out_dir / "password_length_benchmark.xlsx"
    wb.save(path)
    log.info("wrote %s", path)


def run_sweep(
    lengths: list[int],
    *,
    trials: int,
    host: str,
    port: int,
    idc_prefix: str,
    alphabet: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    summaries: list[dict[str, Any]] = []
    all_raw: list[dict[str, Any]] = []

    for L in lengths:
        pw = make_password(L, alphabet=alphabet)
        assert len(pw.encode("utf-8")) == L
        log.info("===== password length=%d =====", L)
        rows = run_trials(
            trials,
            password=pw,
            host=host,
            port=port,
            idc_prefix=f"{idc_prefix}_L{L}",
        )
        for r in rows:
            r["password_length"] = L
        summaries.append(summarize_length(L, rows))
        all_raw.extend(rows)
        s = summaries[-1]
        log.info(
            "L=%d Init=%.2f±%.2f Rec=%.2f±%.2f (%d/%d ok)",
            L,
            s["phases"]["Init≈Enc_proto"]["latency_mean_ms"],
            s["phases"]["Init≈Enc_proto"]["latency_std_ms"],
            s["phases"]["Rec≈Dec_proto"]["latency_mean_ms"],
            s["phases"]["Rec≈Dec_proto"]["latency_std_ms"],
            s["n_success"],
            s["n_trials"],
        )
    return summaries, all_raw


def main() -> None:
    parser = argparse.ArgumentParser(
        description="WBP 实验1：口令长度扫描 → Init/Rec 延迟（对齐 PAEE）"
    )
    parser.add_argument("--trials", type=int, default=20, help="每个口令长度的重复次数")
    parser.add_argument(
        "--lengths",
        default="8,16,32,64,128,256,512",
        help="口令长度列表，或 8:512（2的幂），或 8:512:32（步长）",
    )
    parser.add_argument("--host", default=DEFAULT_HOST)
    parser.add_argument("--port", type=int, default=DEFAULT_PORT)
    parser.add_argument("--alphabet", default="a", help="口令填充用的单 ASCII 字符")
    parser.add_argument(
        "--idc-prefix",
        default="pwlen",
        help="IDC 前缀；每次运行自动追加时间戳，避免远端残留冲突",
    )
    parser.add_argument(
        "--out-dir",
        default=str(ROOT / "experiment" / "output"),
    )
    parser.add_argument("-q", "--quiet", action="store_true")
    args = parser.parse_args()

    level = logging.WARNING if args.quiet else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s %(name)s %(levelname)s %(message)s",
    )
    if args.quiet:
        for name in ("wbp.wire", "wbp.client", "wbp.server", "wbp.hsm", "wbp.bench"):
            logging.getLogger(name).setLevel(logging.WARNING)

    lengths = parse_lengths(args.lengths)
    if any(L < 1 for L in lengths):
        raise SystemExit("password lengths must be >= 1")

    run_tag = datetime.now().strftime("%Y%m%d%H%M%S") + secrets.token_hex(2)
    idc_prefix = f"{args.idc_prefix}_{run_tag}"
    log.info("idc_prefix=%s (unique per run)", idc_prefix)

    host = args.host
    if host in ("0.0.0.0", "::"):
        host = "127.0.0.1"

    check_server(host, args.port)
    summaries, raw = run_sweep(
        lengths,
        trials=args.trials,
        host=host,
        port=args.port,
        idc_prefix=idc_prefix,
        alphabet=args.alphabet,
    )
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    write_outputs(
        summaries,
        raw,
        Path(args.out_dir),
        endpoint=f"{host}:{args.port}",
        trials=args.trials,
        timestamp=ts,
    )

    print()
    header = _length_header(summaries)
    col_w = max(14, max(len(str(h)) for h in header[1:]) + 2)
    print(f"{'【延迟 ms】':<18}" + "".join(f"{h:>{col_w}}" for h in header[1:]))
    print(f"{header[0]:<18}" + "".join(f"{h:>{col_w}}" for h in header[1:]))
    for name, *_rest in PHASES:
        row = _phase_row_mean_std(summaries, name)
        print(f"{row[0]:<18}" + "".join(f"{c:>{col_w}}" for c in row[1:]))
    print(f"\n结果目录: {args.out_dir}")
    print("  password_length_benchmark.xlsx / .md / _summary.csv / .csv")


if __name__ == "__main__":
    main()
