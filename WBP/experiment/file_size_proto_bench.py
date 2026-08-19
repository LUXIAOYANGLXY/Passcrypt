"""实验设计 3（对齐 PAEE file_size_proto_bench）：口令长度固定 16。

测量不同文件大小下：
  1. Init（≈Enc_proto：经 TCP 封装并存储备份密钥 K）
  2. Enc_total = Init + 本地用 K 加密文件（大文件密文**不上传**）
  3. Dec_total = Rec + 本地用恢复的 K' 解密文件

文件大小默认：1, 10, 100, 200, 300, 400, 500 MB。
输出表：行=阶段，列=文件大小。

用法::

    python run_server.py --host 0.0.0.0 --port 8765
    python -m experiment.file_size_proto_bench --trials 3 --host 127.0.0.1 --port 8765
    python -m experiment.file_size_proto_bench --trials 3 --host 54.x.x.x --port 8765 -q
"""

from __future__ import annotations

import argparse
import csv
import logging
import os
import secrets
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from client.app import DEFAULT_HOST, DEFAULT_PORT, Client
from crypto.file_aes import AESGCMCipher
from experiment.tcp_benchmark import _mean_std, check_server

log = logging.getLogger("wbp.filesize")

DISPLAY = "WBP (DFG+23)"
DEFAULT_SIZES_MB = (1, 10, 100, 200, 300, 400, 500)
PASSWORD_LEN = 16


def make_password(length: int = PASSWORD_LEN) -> str:
    return "a" * length


def _gen(size_bytes: int) -> bytes:
    chunk = 8 * 1024 * 1024
    parts: list[bytes] = []
    left = size_bytes
    while left > 0:
        n = min(chunk, left)
        parts.append(os.urandom(n))
        left -= n
    return b"".join(parts)


def run_size_trials(
    *,
    host: str,
    port: int,
    password: str,
    size_mb: int,
    trials: int,
    idc_prefix: str,
    save_ct_dir: Path | None,
) -> list[dict[str, Any]]:
    """
    每个 trial：
      connect → Init（取 K）→ 本地 AES.Enc(file)
      → Rec（取 K'）→ 本地 AES.Dec；大文件密文不上 TCP。
    """
    size = size_mb * 1024 * 1024
    log.info("生成明文 %d MB ...", size_mb)
    plaintext = _gen(size)
    cipher = AESGCMCipher()
    rows: list[dict[str, Any]] = []

    for t in range(1, trials + 1):
        idc = f"{idc_prefix}_{size_mb}mb_{t}"
        log.info("---- %d MB trial %d/%d idc=%s ----", size_mb, t, trials, idc)
        client: Client | None = None
        try:
            client = Client(idc=idc, host=host, port=port, password=password)
            client.connect()

            init_r, init_m = client.init_metrics()
            if not init_r.ok or not init_r.backup_key:
                raise RuntimeError(f"Init 失败: {init_r.error}")
            K = bytes.fromhex(init_r.backup_key)
            init_ms = float(init_m["latency_ms"])
            init_comm = int(init_m["comm_bytes"])

            t0 = time.perf_counter()
            file_ct = cipher.enc(K, plaintext)
            enc_file_ms = (time.perf_counter() - t0) * 1000.0
            ct_bytes = len(file_ct.serialize())

            if save_ct_dir is not None:
                save_ct_dir.mkdir(parents=True, exist_ok=True)
                (save_ct_dir / f"{idc}.ct").write_bytes(file_ct.serialize())

            enc_total_ms = init_ms + enc_file_ms

            rec_r, rec_m = client.recover_metrics()
            if not rec_r.ok or not rec_r.backup_key:
                raise RuntimeError(f"Rec 失败: {rec_r.error}")
            K2 = bytes.fromhex(rec_r.backup_key)
            if K2 != K:
                raise RuntimeError("Init/Rec 密钥不一致")
            rec_ms = float(rec_m["latency_ms"])
            rec_comm = int(rec_m["comm_bytes"])

            t1 = time.perf_counter()
            out = cipher.dec(K2, file_ct)
            dec_file_ms = (time.perf_counter() - t1) * 1000.0
            if out != plaintext:
                raise RuntimeError("plaintext mismatch")

            dec_total_ms = rec_ms + dec_file_ms

            rows.append(
                {
                    "success": True,
                    "size_mb": size_mb,
                    "trial_id": t,
                    "idc": idc,
                    "init_ms": init_ms,
                    "enc_file_ms": enc_file_ms,
                    "enc_total_ms": enc_total_ms,
                    "rec_ms": rec_ms,
                    "dec_file_ms": dec_file_ms,
                    "dec_total_ms": dec_total_ms,
                    "init_comm_bytes": init_comm,
                    "rec_comm_bytes": rec_comm,
                    "ct_bytes": ct_bytes,
                    "error": "",
                }
            )
            log.info(
                "  Init=%.1f Enc_file=%.1f Enc_total=%.1f "
                "Rec=%.1f Dec_file=%.1f Dec_total=%.1f ct=%dB",
                init_ms,
                enc_file_ms,
                enc_total_ms,
                rec_ms,
                dec_file_ms,
                dec_total_ms,
                ct_bytes,
            )
            del file_ct, out
        except Exception as e:
            log.error("trial fail: %s", e)
            rows.append(
                {
                    "success": False,
                    "size_mb": size_mb,
                    "trial_id": t,
                    "idc": idc,
                    "init_ms": 0.0,
                    "enc_file_ms": 0.0,
                    "enc_total_ms": 0.0,
                    "rec_ms": 0.0,
                    "dec_file_ms": 0.0,
                    "dec_total_ms": 0.0,
                    "init_comm_bytes": 0,
                    "rec_comm_bytes": 0,
                    "ct_bytes": 0,
                    "error": str(e),
                }
            )
        finally:
            if client is not None:
                client.close()

    del plaintext
    return rows


# 主表（对齐 PAEE：无单独 Ext；用 Init 对标协议侧）
PHASE_ROWS = (
    ("Init≈Enc_proto", "init_ms"),
    ("Enc（Init+本地加密文件）", "enc_total_ms"),
    ("Dec（Rec+本地解密文件）", "dec_total_ms"),
)

BREAKDOWN_ROWS = (
    ("Init≈Enc_proto", "init_ms"),
    ("Enc_file（本地 AES-GCM）", "enc_file_ms"),
    ("Enc_total（Init+file）", "enc_total_ms"),
    ("Rec≈Dec_proto", "rec_ms"),
    ("Dec_file（本地 AES-GCM）", "dec_file_ms"),
    ("Dec_total（Rec+file）", "dec_total_ms"),
)


def summarize_by_size(
    all_raw: list[dict[str, Any]], sizes: list[int]
) -> dict[int, dict[str, Any]]:
    out: dict[int, dict[str, Any]] = {}
    for mb in sizes:
        rows = [r for r in all_raw if r["size_mb"] == mb]
        ok = [r for r in rows if r["success"]]
        stats: dict[str, Any] = {
            "n_trials": len(rows),
            "n_success": len(ok),
        }
        for _name, key in BREAKDOWN_ROWS:
            m, s = _mean_std([float(r[key]) for r in ok])
            stats[f"{key}_mean"] = m
            stats[f"{key}_std"] = s
        out[mb] = stats
    return out


def _size_header(sizes: list[int]) -> list[str]:
    return ["阶段"] + [f"{mb}MB" for mb in sizes]


def _row_mean_std(
    by_size: dict[int, dict[str, Any]],
    sizes: list[int],
    phase: str,
    key: str,
) -> list[Any]:
    cells: list[Any] = [phase]
    for mb in sizes:
        st = by_size[mb]
        cells.append(f"{st[f'{key}_mean']:.2f} ± {st[f'{key}_std']:.2f}")
    return cells


def write_outputs(
    by_size: dict[int, dict[str, Any]],
    raw: list[dict[str, Any]],
    sizes: list[int],
    out_dir: Path,
    *,
    endpoint: str,
    trials: int,
    password_len: int,
    timestamp: str,
) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    note = (
        f"实验3：口令长度={password_len}；trials/大小={trials}；Server={endpoint}；"
        "Enc_total=Init+本地加密文件（大文件密文不上传）；"
        "Dec=Rec+本地解密；列=文件大小"
    )
    header = _size_header(sizes)
    sep = "|" + "|".join(["------"] * len(header)) + "|"

    md = [
        "# WBP 文件大小：Init / Enc / Dec 延迟",
        "",
        f"生成时间: {timestamp}",
        note,
        "",
        "## 主表 mean±std (ms)",
        "",
        "| " + " | ".join(header) + " |",
        sep,
    ]
    for phase, key in PHASE_ROWS:
        md.append(
            "| "
            + " | ".join(str(c) for c in _row_mean_std(by_size, sizes, phase, key))
            + " |"
        )
    md += ["", "## 分解 mean±std (ms)", "", "| " + " | ".join(header) + " |", sep]
    for phase, key in BREAKDOWN_ROWS:
        md.append(
            "| "
            + " | ".join(str(c) for c in _row_mean_std(by_size, sizes, phase, key))
            + " |"
        )
    (out_dir / "file_size_proto_benchmark.md").write_text(
        "\n".join(md) + "\n", encoding="utf-8"
    )

    with (out_dir / "file_size_proto_benchmark_summary.csv").open(
        "w", newline="", encoding="utf-8"
    ) as f:
        w = csv.writer(f)
        w.writerow(header)
        for phase, key in PHASE_ROWS:
            w.writerow(_row_mean_std(by_size, sizes, phase, key))
        w.writerow([])
        w.writerow(["分解"])
        for phase, key in BREAKDOWN_ROWS:
            w.writerow(_row_mean_std(by_size, sizes, phase, key))

    if raw:
        with (out_dir / "file_size_proto_benchmark.csv").open(
            "w", newline="", encoding="utf-8"
        ) as f:
            w = csv.DictWriter(f, fieldnames=list(raw[0].keys()))
            w.writeheader()
            w.writerows(raw)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        log.warning("openpyxl missing; skip xlsx")
        return

    def style_h(ws) -> None:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"
    ws.append(header)
    style_h(ws)
    for phase, key in PHASE_ROWS:
        ws.append(_row_mean_std(by_size, sizes, phase, key))
    ws.append([])
    ws.append([note])
    ws.append([f"方案={DISPLAY}", f"时间={timestamp}"])

    ws2 = wb.create_sheet("分解")
    ws2.append(header)
    style_h(ws2)
    for phase, key in BREAKDOWN_ROWS:
        ws2.append(_row_mean_std(by_size, sizes, phase, key))

    ws3 = wb.create_sheet("延迟_mean")
    ws3.append(header)
    style_h(ws3)
    for phase, key in PHASE_ROWS:
        ws3.append(
            [phase] + [round(by_size[mb][f"{key}_mean"], 3) for mb in sizes]
        )

    ws4 = wb.create_sheet("原始数据")
    if raw:
        keys = list(raw[0].keys())
        ws4.append(keys)
        style_h(ws4)
        for r in raw:
            ws4.append([r[k] for k in keys])

    path = out_dir / "file_size_proto_benchmark.xlsx"
    wb.save(path)
    log.info("wrote %s", path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="WBP 实验3：文件大小 vs Enc/Dec（口令长度=16，对齐 PAEE）"
    )
    parser.add_argument("--trials", type=int, default=3)
    parser.add_argument(
        "--sizes",
        type=int,
        nargs="+",
        default=list(DEFAULT_SIZES_MB),
        help="文件大小 MB，默认 1 10 100 200 300 400 500",
    )
    parser.add_argument("--password-len", type=int, default=PASSWORD_LEN)
    parser.add_argument("--host", default=DEFAULT_HOST)
    parser.add_argument("--port", type=int, default=DEFAULT_PORT)
    parser.add_argument(
        "--save-ct",
        action="store_true",
        help="将文件密文写到 experiment/output/file_cts/（不计时）",
    )
    parser.add_argument(
        "--out-dir", default=str(ROOT / "experiment" / "output")
    )
    parser.add_argument("-q", "--quiet", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(asctime)s %(name)s %(levelname)s %(message)s",
    )
    if args.quiet:
        for name in ("wbp.wire", "wbp.client", "wbp.server", "wbp.hsm", "wbp.bench"):
            logging.getLogger(name).setLevel(logging.WARNING)

    password = make_password(args.password_len)
    assert len(password) == args.password_len

    host = args.host
    if host in ("0.0.0.0", "::"):
        host = "127.0.0.1"
    check_server(host, args.port)

    run_tag = datetime.now().strftime("%Y%m%d%H%M%S") + secrets.token_hex(2)
    idc_prefix = f"fsize_{run_tag}"
    save_ct_dir = (
        Path(args.out_dir) / "file_cts" / run_tag if args.save_ct else None
    )

    all_raw: list[dict[str, Any]] = []
    for mb in args.sizes:
        log.info("===== file size %d MB =====", mb)
        all_raw.extend(
            run_size_trials(
                host=host,
                port=args.port,
                password=password,
                size_mb=mb,
                trials=args.trials,
                idc_prefix=idc_prefix,
                save_ct_dir=save_ct_dir,
            )
        )

    by_size = summarize_by_size(all_raw, args.sizes)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    write_outputs(
        by_size,
        all_raw,
        args.sizes,
        Path(args.out_dir),
        endpoint=f"{host}:{args.port}",
        trials=args.trials,
        password_len=args.password_len,
        timestamp=ts,
    )

    header = _size_header(args.sizes)
    col_w = max(16, max(len(h) for h in header[1:]) + 2)
    print()
    print(f"{header[0]:<42}" + "".join(f"{h:>{col_w}}" for h in header[1:]))
    for phase, key in PHASE_ROWS:
        row = _row_mean_std(by_size, args.sizes, phase, key)
        print(f"{row[0]:<42}" + "".join(f"{c:>{col_w}}" for c in row[1:]))
    print(f"\n结果: {args.out_dir}/file_size_proto_benchmark.xlsx|.md|.csv")


if __name__ == "__main__":
    main()
