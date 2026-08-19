"""使用经 TCP 获得的 WBP backup key K，对大文件做 AES-GCM 加密/解密并测时。

对齐 PPKRv1 ``experiment/file_encrypt_bench.py``。

用法::

    python run_server.py --port 8876
    python -m experiment.file_encrypt_bench --port 8876 --trials 5

输出::
    experiment/output/file_encrypt_benchmark.xlsx
    experiment/output/file_decrypt_benchmark.xlsx
"""

from __future__ import annotations

import argparse
import logging
import os
import socket
import statistics
import sys
import time
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from client.app import DEFAULT_HOST, DEFAULT_PORT, Client
from crypto.file_aes import AESCiphertext, AESGCMCipher

log = logging.getLogger("wbp.file-enc")

DEFAULT_SIZES_MB = (1, 10, 100, 200, 300, 400, 500)
PROTOCOL_DISPLAY = "WBP (DFG+23)"


def _mean_std(values: list[float]) -> tuple[float, float]:
    if not values:
        return 0.0, 0.0
    if len(values) == 1:
        return values[0], 0.0
    return statistics.mean(values), statistics.stdev(values)


def _throughput_mbps(size_mb: int, ms: float) -> float:
    return (size_mb / (ms / 1000.0)) if ms > 0 else 0.0


def check_server(host: str, port: int) -> None:
    try:
        with socket.create_connection((host, port), timeout=3):
            pass
    except OSError as e:
        raise SystemExit(
            f"无法连接 WBP Server {host}:{port}，请先启动: python run_server.py\n原因: {e}"
        ) from e


def obtain_keys_via_tcp(
    password: str, idc: str, host: str, port: int
) -> tuple[bytes, bytes, dict, dict]:
    """经 TCP Init/Rec 得到 K 与 K'（含网络）。"""
    log.info("TCP 获取密钥 idc=%s %s:%d", idc, host, port)
    client = Client(idc=idc, host=host, port=port, password=password)
    try:
        client.connect()
        init_r, init_m = client.init_metrics()
        if not init_r.ok or not init_r.backup_key:
            raise RuntimeError(f"Init 失败: {init_r.error}")
        rec_r, rec_m = client.recover_metrics()
        if not rec_r.ok or not rec_r.backup_key:
            raise RuntimeError(f"Rec 失败: {rec_r.error}")
        K = bytes.fromhex(init_r.backup_key)
        K_rec = bytes.fromhex(rec_r.backup_key)
        if K != K_rec:
            raise RuntimeError("Init/Rec 密钥不一致")
        log.info(
            "TCP Init=%.2fms/%dB Rec=%.2fms/%dB K=%s...",
            init_m["latency_ms"],
            init_m["comm_bytes"],
            rec_m["latency_ms"],
            rec_m["comm_bytes"],
            K.hex()[:16],
        )
        return K, K_rec, init_m, rec_m
    finally:
        client.close()


def generate_plaintext(size_bytes: int, path: Path | None) -> bytes:
    log.info("生成明文 %.0f MB ...", size_bytes / (1024 * 1024))
    chunk = 8 * 1024 * 1024
    parts: list[bytes] = []
    remaining = size_bytes
    while remaining > 0:
        n = min(chunk, remaining)
        parts.append(os.urandom(n))
        remaining -= n
    data = b"".join(parts)
    if path is not None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(data)
        log.info("明文已写入 %s", path)
    return data


def encrypt_once(
    cipher: AESGCMCipher, key: bytes, plaintext: bytes
) -> tuple[float, AESCiphertext, int]:
    t0 = time.perf_counter()
    ct = cipher.enc(key, plaintext)
    elapsed_ms = (time.perf_counter() - t0) * 1000
    return elapsed_ms, ct, len(ct.serialize())


def decrypt_once(
    cipher: AESGCMCipher, key: bytes, ct: AESCiphertext
) -> tuple[float, bytes | None]:
    t0 = time.perf_counter()
    pt = cipher.dec(key, ct)
    elapsed_ms = (time.perf_counter() - t0) * 1000
    return elapsed_ms, pt


def run_size_benchmark(
    cipher: AESGCMCipher,
    key_enc: bytes,
    key_dec: bytes,
    size_mb: int,
    trials: int,
    data_dir: Path | None,
    keep_files: bool,
) -> dict:
    size_bytes = size_mb * 1024 * 1024
    plain_path = (data_dir / f"plain_{size_mb}mb.bin") if data_dir and keep_files else None
    plaintext = generate_plaintext(size_bytes, plain_path)

    enc_times_ms: list[float] = []
    dec_times_ms: list[float] = []
    ct_lens: list[int] = []
    fixed_ct: AESCiphertext | None = None

    log.info("  [%d MB] 加密测时 trials=%d (key=Init K)", size_mb, trials)
    for i in range(1, trials + 1):
        ms, ct, ct_len = encrypt_once(cipher, key_enc, plaintext)
        enc_times_ms.append(ms)
        ct_lens.append(ct_len)
        if fixed_ct is None:
            fixed_ct = ct
        log.info("  encrypt %d MB %d/%d: %.2f ms", size_mb, i, trials, ms)

    assert fixed_ct is not None

    if keep_files and data_dir is not None:
        out = data_dir / f"cipher_{size_mb}mb.bin"
        out.write_bytes(fixed_ct.serialize())

    check_pt = cipher.dec(key_dec, fixed_ct)
    if check_pt != plaintext:
        raise RuntimeError(f"{size_mb} MB 用恢复密钥解密校验失败")
    del plaintext
    del check_pt

    log.info("  [%d MB] 解密测时 trials=%d (key=Rec K')", size_mb, trials)
    for i in range(1, trials + 1):
        ms, pt = decrypt_once(cipher, key_dec, fixed_ct)
        if pt is None:
            raise RuntimeError(f"{size_mb} MB 第 {i} 次解密失败")
        dec_times_ms.append(ms)
        log.info("  decrypt %d MB %d/%d: %.2f ms", size_mb, i, trials, ms)
        del pt

    enc_mean, enc_std = _mean_std(enc_times_ms)
    dec_mean, dec_std = _mean_std(dec_times_ms)
    return {
        "size_mb": size_mb,
        "size_bytes": size_bytes,
        "trials": trials,
        "encrypt_mean_ms": enc_mean,
        "encrypt_std_ms": enc_std,
        "encrypt_throughput_mean_MBps": statistics.mean(
            [_throughput_mbps(size_mb, t) for t in enc_times_ms]
        ),
        "decrypt_mean_ms": dec_mean,
        "decrypt_std_ms": dec_std,
        "decrypt_throughput_mean_MBps": statistics.mean(
            [_throughput_mbps(size_mb, t) for t in dec_times_ms]
        ),
        "ciphertext_bytes_mean": statistics.mean(ct_lens) if ct_lens else 0,
        "raw_encrypt_ms": enc_times_ms,
        "raw_decrypt_ms": dec_times_ms,
    }


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = col[0].column_letter
        width = min(48, max(12, max(len(str(c.value or "")) for c in col) + 2))
        ws.column_dimensions[letter].width = width


def write_encrypt_excel(
    summary_rows: list[dict],
    raw_rows: list[dict],
    output_path: Path,
    meta: dict,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "加密汇总"
    headers = [
        "协议",
        "密钥K(hex前16)",
        "文件大小(MB)",
        "文件大小(bytes)",
        "试验次数",
        "加密时间_mean(ms)",
        "加密时间_std(ms)",
        "加密吞吐_mean(MB/s)",
        "密文大小_mean(bytes)",
        "AE算法",
        "时间戳",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in summary_rows:
        ws.append(
            [
                meta["protocol"],
                meta["key_prefix"],
                row["size_mb"],
                row["size_bytes"],
                row["trials"],
                round(row["encrypt_mean_ms"], 3),
                round(row["encrypt_std_ms"], 3),
                round(row["encrypt_throughput_mean_MBps"], 3),
                int(row["ciphertext_bytes_mean"]),
                "AES-256-GCM (crypto.file_aes.AESGCMCipher)",
                meta["timestamp"],
            ]
        )

    ws2 = wb.create_sheet("加密原始数据")
    ws2.append(["协议", "文件大小(MB)", "trial_id", "加密时间(ms)", "吞吐(MB/s)"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for r in raw_rows:
        ws2.append(
            [
                meta["protocol"],
                r["size_mb"],
                r["trial_id"],
                round(r["encrypt_ms"], 3),
                round(r["encrypt_throughput_MBps"], 3),
            ]
        )
    _autosize(ws)
    _autosize(ws2)
    wb.save(output_path)
    return output_path


def write_decrypt_excel(
    summary_rows: list[dict],
    raw_rows: list[dict],
    output_path: Path,
    meta: dict,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "解密汇总"
    headers = [
        "协议",
        "密钥K(hex前16)",
        "文件大小(MB)",
        "文件大小(bytes)",
        "试验次数",
        "解密时间_mean(ms)",
        "解密时间_std(ms)",
        "解密吞吐_mean(MB/s)",
        "密文大小_mean(bytes)",
        "AE算法",
        "时间戳",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in summary_rows:
        ws.append(
            [
                meta["protocol"],
                meta["key_prefix"],
                row["size_mb"],
                row["size_bytes"],
                row["trials"],
                round(row["decrypt_mean_ms"], 3),
                round(row["decrypt_std_ms"], 3),
                round(row["decrypt_throughput_mean_MBps"], 3),
                int(row["ciphertext_bytes_mean"]),
                "AES-256-GCM (crypto.file_aes.AESGCMCipher)",
                meta["timestamp"],
            ]
        )

    ws2 = wb.create_sheet("解密原始数据")
    ws2.append(["协议", "文件大小(MB)", "trial_id", "解密时间(ms)", "吞吐(MB/s)"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for r in raw_rows:
        ws2.append(
            [
                meta["protocol"],
                r["size_mb"],
                r["trial_id"],
                round(r["decrypt_ms"], 3),
                round(r["decrypt_throughput_MBps"], 3),
            ]
        )
    _autosize(ws)
    _autosize(ws2)
    wb.save(output_path)
    return output_path


def print_summary(summary_rows: list[dict], enc_path: Path, dec_path: Path, meta: dict) -> None:
    print("\n" + "=" * 78)
    print("  数据密钥 K — 大文件加密/解密（密钥经 TCP Init/Rec 获得）")
    print("=" * 78)
    print(f"  协议: {meta['protocol']}  端点: {meta.get('endpoint', '')}")
    print(f"  K={meta['key_prefix']}...  AE: AES-256-GCM")
    print(
        f"  TCP Init: {meta.get('init_ms', 0):.2f} ms / {meta.get('init_comm', 0)} B"
        f"  Rec: {meta.get('rec_ms', 0):.2f} ms / {meta.get('rec_comm', 0)} B"
    )
    print("-" * 78)
    print(
        f"  {'大小(MB)':>8}  {'加密mean':>10}  {'加密std':>8}  "
        f"{'解密mean':>10}  {'解密std':>8}  {'解密吞吐':>10}"
    )
    for row in summary_rows:
        print(
            f"  {row['size_mb']:>8}  {row['encrypt_mean_ms']:>10.2f}  {row['encrypt_std_ms']:>8.2f}  "
            f"{row['decrypt_mean_ms']:>10.2f}  {row['decrypt_std_ms']:>8.2f}  "
            f"{row['decrypt_throughput_mean_MBps']:>10.2f}"
        )
    print("-" * 78)
    print(f"  加密 Excel: {enc_path}")
    print(f"  解密 Excel: {dec_path}")
    print("=" * 78 + "\n")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="TCP 获取 WBP K 后加密/解密大文件并测时 → Excel"
    )
    parser.add_argument("--trials", type=int, default=5)
    parser.add_argument("--password", default="file_enc_bench_pw")
    parser.add_argument("--idc", default="wbp_file_user")
    parser.add_argument("--host", default=DEFAULT_HOST)
    parser.add_argument("--port", type=int, default=DEFAULT_PORT)
    parser.add_argument("--sizes", type=int, nargs="+", default=list(DEFAULT_SIZES_MB))
    parser.add_argument(
        "--encrypt-output",
        type=str,
        default=str(ROOT / "experiment" / "output" / "file_encrypt_benchmark.xlsx"),
    )
    parser.add_argument(
        "--decrypt-output",
        type=str,
        default=str(ROOT / "experiment" / "output" / "file_decrypt_benchmark.xlsx"),
    )
    parser.add_argument("--keep-files", action="store_true")
    parser.add_argument("-q", "--quiet", action="store_true")
    args = parser.parse_args()

    level = logging.WARNING if args.quiet else logging.INFO
    logging.basicConfig(level=level, format="%(asctime)s %(name)s %(levelname)s %(message)s")
    if args.quiet:
        for name in ("wbp.wire", "wbp.client", "wbp.server", "wbp.hsm"):
            logging.getLogger(name).setLevel(logging.WARNING)

    check_server(args.host, args.port)
    data_dir = ROOT / "experiment" / "output" / "file_encrypt"
    if args.keep_files:
        data_dir.mkdir(parents=True, exist_ok=True)

    K, K_rec, init_m, rec_m = obtain_keys_via_tcp(
        args.password, args.idc, args.host, args.port
    )
    cipher = AESGCMCipher()

    sample = os.urandom(1024)
    ct = cipher.enc(K, sample)
    pt = cipher.dec(K_rec, ct)
    if pt != sample:
        raise RuntimeError("AES-GCM 加解密自检失败")

    summary_rows: list[dict] = []
    enc_raw: list[dict] = []
    dec_raw: list[dict] = []

    for size_mb in args.sizes:
        log.info("======== 开始 %d MB（加密 + 解密）========", size_mb)
        result = run_size_benchmark(
            cipher,
            K,
            K_rec,
            size_mb,
            args.trials,
            data_dir if args.keep_files else None,
            args.keep_files,
        )
        summary_rows.append(result)
        for i, ms in enumerate(result["raw_encrypt_ms"], start=1):
            enc_raw.append(
                {
                    "size_mb": size_mb,
                    "trial_id": i,
                    "encrypt_ms": ms,
                    "encrypt_throughput_MBps": _throughput_mbps(size_mb, ms),
                }
            )
        for i, ms in enumerate(result["raw_decrypt_ms"], start=1):
            dec_raw.append(
                {
                    "size_mb": size_mb,
                    "trial_id": i,
                    "decrypt_ms": ms,
                    "decrypt_throughput_MBps": _throughput_mbps(size_mb, ms),
                }
            )

    meta = {
        "protocol": PROTOCOL_DISPLAY,
        "key_prefix": K.hex()[:16],
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "endpoint": f"{args.host}:{args.port}",
        "init_ms": init_m["latency_ms"],
        "init_comm": init_m["comm_bytes"],
        "rec_ms": rec_m["latency_ms"],
        "rec_comm": rec_m["comm_bytes"],
    }
    enc_path = write_encrypt_excel(summary_rows, enc_raw, Path(args.encrypt_output), meta)
    dec_path = write_decrypt_excel(summary_rows, dec_raw, Path(args.decrypt_output), meta)
    print_summary(summary_rows, enc_path, dec_path, meta)


if __name__ == "__main__":
    main()
