"""端到端一次性测量（含 TCP）：Init → 加密 → Rec → 解密。

对齐 PPKRv1 ``experiment/e2e_key_file_bench.py``。

测量项（多次 trial 取平均）::
    1. Init 时间（含 TCP，生成 backup key K）
    2. 用 Init 的 K 加密各大小文件的时间
    3. Rec 时间（含 TCP，恢复密钥 K'）
    4. 用恢复的 K' 解密对应密文的时间

用法::

    python run_server.py --port 8876
    python -m experiment.e2e_key_file_bench --port 8876 --trials 5

输出::
    experiment/output/e2e_key_file_benchmark.xlsx
"""

from __future__ import annotations

import argparse
import logging
import os
import socket
import statistics
import sys
import time
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from client.app import DEFAULT_HOST, DEFAULT_PORT, Client
from crypto.file_aes import AESCiphertext, AESGCMCipher

log = logging.getLogger("wbp.e2e")

DEFAULT_SIZES_MB = (1, 10, 100, 200, 300, 400, 500)
PROTOCOL_DISPLAY = "WBP (DFG+23)"


def _mean_std(values: list[float]) -> tuple[float, float]:
    if not values:
        return 0.0, 0.0
    if len(values) == 1:
        return values[0], 0.0
    return statistics.mean(values), statistics.stdev(values)


def _tp(size_mb: int, ms: float) -> float:
    return (size_mb / (ms / 1000.0)) if ms > 0 else 0.0


def _gen_plain(size_bytes: int) -> bytes:
    chunk = 8 * 1024 * 1024
    parts: list[bytes] = []
    left = size_bytes
    while left > 0:
        n = min(chunk, left)
        parts.append(os.urandom(n))
        left -= n
    return b"".join(parts)


def check_server(host: str, port: int) -> None:
    try:
        with socket.create_connection((host, port), timeout=3):
            pass
    except OSError as e:
        raise SystemExit(
            f"无法连接 WBP Server {host}:{port}，请先启动: python run_server.py\n原因: {e}"
        ) from e


def run_one_trial(
    trial_id: int,
    password: str,
    sizes_mb: list[int],
    host: str,
    port: int,
) -> dict:
    """含网络的完整流水线：TCP Init → 加密 → TCP Rec → 解密。"""
    idc = f"e2e_wbp_{trial_id}"
    cipher = AESGCMCipher()
    client = Client(idc=idc, host=host, port=port, password=password)

    log.info("==== WBP trial %d idc=%s %s:%d ====", trial_id, idc, host, port)

    try:
        client.connect()

        init_r, init_m = client.init_metrics()
        if not init_r.ok or not init_r.backup_key:
            raise RuntimeError(f"Init 失败: {init_r.error}")
        K = bytes.fromhex(init_r.backup_key)
        log.info(
            "  Init(TCP): %.2f ms  comm=%d B  K=%s...",
            init_m["latency_ms"],
            init_m["comm_bytes"],
            K.hex()[:16],
        )

        ciphertexts: dict[int, AESCiphertext] = {}
        enc_times: dict[int, float] = {}
        for size_mb in sizes_mb:
            plain = _gen_plain(size_mb * 1024 * 1024)
            t0 = time.perf_counter()
            ct = cipher.enc(K, plain)
            enc_ms = (time.perf_counter() - t0) * 1000
            check = cipher.dec(K, ct)
            if check != plain:
                raise RuntimeError(f"{size_mb} MB 加密后自检失败")
            del plain, check
            ciphertexts[size_mb] = ct
            enc_times[size_mb] = enc_ms
            log.info("  Encrypt %d MB: %.2f ms", size_mb, enc_ms)

        rec_r, rec_m = client.recover_metrics()
        if not rec_r.ok or not rec_r.backup_key:
            raise RuntimeError(f"Rec 失败: {rec_r.error}")
        K_rec = bytes.fromhex(rec_r.backup_key)
        if K_rec != K:
            raise RuntimeError("恢复密钥与 Init 密钥不一致")
        log.info(
            "  Rec(TCP): %.2f ms  comm=%d B  K_rec 一致",
            rec_m["latency_ms"],
            rec_m["comm_bytes"],
        )

        dec_times: dict[int, float] = {}
        for size_mb in sizes_mb:
            ct = ciphertexts[size_mb]
            t0 = time.perf_counter()
            pt = cipher.dec(K_rec, ct)
            dec_ms = (time.perf_counter() - t0) * 1000
            if pt is None or len(pt) != size_mb * 1024 * 1024:
                raise RuntimeError(f"{size_mb} MB 用恢复密钥解密失败")
            del pt
            dec_times[size_mb] = dec_ms
            log.info("  Decrypt %d MB (recovered K): %.2f ms", size_mb, dec_ms)

        return {
            "protocol": "wbp",
            "display": PROTOCOL_DISPLAY,
            "trial_id": trial_id,
            "idc": idc,
            "init_ms": init_m["latency_ms"],
            "rec_ms": rec_m["latency_ms"],
            "init_comm_bytes": init_m["comm_bytes"],
            "rec_comm_bytes": rec_m["comm_bytes"],
            "key_match": True,
            "enc_times": enc_times,
            "dec_times": dec_times,
            "success": True,
            "error": "",
        }
    finally:
        client.close()


def aggregate_results(trials: list[dict], sizes_mb: list[int]) -> dict:
    ok = [t for t in trials if t["success"]]
    init_ms = [t["init_ms"] for t in ok]
    rec_ms = [t["rec_ms"] for t in ok]
    init_comm = [float(t["init_comm_bytes"]) for t in ok]
    rec_comm = [float(t["rec_comm_bytes"]) for t in ok]
    init_mean, init_std = _mean_std(init_ms)
    rec_mean, rec_std = _mean_std(rec_ms)
    init_cm, init_cs = _mean_std(init_comm)
    rec_cm, rec_cs = _mean_std(rec_comm)

    file_rows = []
    for size_mb in sizes_mb:
        enc = [t["enc_times"][size_mb] for t in ok]
        dec = [t["dec_times"][size_mb] for t in ok]
        em, es = _mean_std(enc)
        dm, ds = _mean_std(dec)
        file_rows.append(
            {
                "size_mb": size_mb,
                "encrypt_mean_ms": em,
                "encrypt_std_ms": es,
                "encrypt_tp_mean": statistics.mean([_tp(size_mb, x) for x in enc]) if enc else 0,
                "decrypt_mean_ms": dm,
                "decrypt_std_ms": ds,
                "decrypt_tp_mean": statistics.mean([_tp(size_mb, x) for x in dec]) if dec else 0,
            }
        )

    return {
        "protocol": "wbp",
        "display": PROTOCOL_DISPLAY,
        "n_trials": len(trials),
        "n_success": len(ok),
        "init_mean_ms": init_mean,
        "init_std_ms": init_std,
        "rec_mean_ms": rec_mean,
        "rec_std_ms": rec_std,
        "init_comm_mean": init_cm,
        "init_comm_std": init_cs,
        "rec_comm_mean": rec_cm,
        "rec_comm_std": rec_cs,
        "file_rows": file_rows,
    }


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = col[0].column_letter
        width = min(44, max(11, max(len(str(c.value or "")) for c in col) + 2))
        ws.column_dimensions[letter].width = width


def write_excel(
    aggregates: list[dict],
    raw_trials: list[dict],
    sizes_mb: list[int],
    output_path: Path,
    meta: dict,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "协议阶段"
    ws1.append(
        [
            "方案",
            "试验次数",
            "成功次数",
            "Init_mean(ms)",
            "Init_std(ms)",
            "Init通信_mean(bytes)",
            "Rec_mean(ms)",
            "Rec_std(ms)",
            "Rec通信_mean(bytes)",
            "Server 端点",
            "说明",
            "时间戳",
        ]
    )
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    note = "Init/Rec 含 TCP 网络；加密用 Init 的 K，解密用 Rec 恢复的 K'"
    for agg in aggregates:
        ws1.append(
            [
                agg["display"],
                agg["n_trials"],
                agg["n_success"],
                round(agg["init_mean_ms"], 3),
                round(agg["init_std_ms"], 3),
                round(agg["init_comm_mean"], 1),
                round(agg["rec_mean_ms"], 3),
                round(agg["rec_std_ms"], 3),
                round(agg["rec_comm_mean"], 1),
                meta["url"],
                note,
                meta["timestamp"],
            ]
        )

    ws2 = wb.create_sheet("文件加解密")
    ws2.append(
        [
            "方案",
            "文件大小(MB)",
            "加密_mean(ms)",
            "加密_std(ms)",
            "加密吞吐(MB/s)",
            "解密_mean(ms)",
            "解密_std(ms)",
            "解密吞吐(MB/s)",
            "密钥来源",
            "时间戳",
        ]
    )
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for agg in aggregates:
        for row in agg["file_rows"]:
            ws2.append(
                [
                    agg["display"],
                    row["size_mb"],
                    round(row["encrypt_mean_ms"], 3),
                    round(row["encrypt_std_ms"], 3),
                    round(row["encrypt_tp_mean"], 3),
                    round(row["decrypt_mean_ms"], 3),
                    round(row["decrypt_std_ms"], 3),
                    round(row["decrypt_tp_mean"], 3),
                    "K 经 TCP Init 获得；K' 经 TCP Rec 恢复",
                    meta["timestamp"],
                ]
            )

    ws3 = wb.create_sheet("原始数据")
    headers = [
        "方案",
        "trial_id",
        "idc",
        "success",
        "Init(ms)",
        "Init通信(bytes)",
        "Rec(ms)",
        "Rec通信(bytes)",
        "key_match",
    ]
    for s in sizes_mb:
        headers += [f"加密_{s}MB(ms)", f"解密_{s}MB(ms)"]
    headers.append("error")
    ws3.append(headers)
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    for t in raw_trials:
        row = [
            t["display"],
            t["trial_id"],
            t["idc"],
            t["success"],
            round(t.get("init_ms", 0), 3),
            t.get("init_comm_bytes", 0),
            round(t.get("rec_ms", 0), 3),
            t.get("rec_comm_bytes", 0),
            t.get("key_match", False),
        ]
        for s in sizes_mb:
            row.append(round(t.get("enc_times", {}).get(s, 0), 3))
            row.append(round(t.get("dec_times", {}).get(s, 0), 3))
        row.append(t.get("error", ""))
        ws3.append(row)

    for ws in (ws1, ws2, ws3):
        _autosize(ws)
    wb.save(output_path)
    return output_path


def print_summary(aggregates: list[dict], path: Path, endpoint: str) -> None:
    print("\n" + "=" * 78)
    print("  端到端（含 TCP）：Init → 加密 → Rec → 解密")
    print(f"  Server: {endpoint}")
    print("=" * 78)
    for agg in aggregates:
        print(f"\n【{agg['display']}】 n={agg['n_trials']} 成功={agg['n_success']}")
        print(
            f"  Init(TCP): {agg['init_mean_ms']:.2f} ± {agg['init_std_ms']:.2f} ms"
            f"  通信 {agg['init_comm_mean']:.0f} B"
        )
        print(
            f"  Rec (TCP): {agg['rec_mean_ms']:.2f} ± {agg['rec_std_ms']:.2f} ms"
            f"  通信 {agg['rec_comm_mean']:.0f} B"
        )
        print(
            f"  {'大小MB':>8}  {'加密mean':>10}  {'加密std':>8}  "
            f"{'解密mean':>10}  {'解密std':>8}"
        )
        for row in agg["file_rows"]:
            print(
                f"  {row['size_mb']:>8}  {row['encrypt_mean_ms']:>10.2f}  "
                f"{row['encrypt_std_ms']:>8.2f}  {row['decrypt_mean_ms']:>10.2f}  "
                f"{row['decrypt_std_ms']:>8.2f}"
            )
    print("-" * 78)
    print(f"  Excel: {path}")
    print("=" * 78 + "\n")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="端到端含网络：TCP Init/加密/TCP Rec/解密 → Excel"
    )
    parser.add_argument("--trials", type=int, default=5)
    parser.add_argument("--password", default="e2e_key_file_pw")
    parser.add_argument("--host", default=DEFAULT_HOST)
    parser.add_argument("--port", type=int, default=DEFAULT_PORT)
    parser.add_argument("--sizes", type=int, nargs="+", default=list(DEFAULT_SIZES_MB))
    parser.add_argument(
        "--output",
        type=str,
        default=str(ROOT / "experiment" / "output" / "e2e_key_file_benchmark.xlsx"),
    )
    parser.add_argument("-q", "--quiet", action="store_true")
    args = parser.parse_args()

    level = logging.WARNING if args.quiet else logging.INFO
    logging.basicConfig(level=level, format="%(asctime)s %(name)s %(levelname)s %(message)s")
    if args.quiet:
        for name in ("wbp.wire", "wbp.client", "wbp.server", "wbp.hsm"):
            logging.getLogger(name).setLevel(logging.WARNING)

    check_server(args.host, args.port)
    endpoint = f"{args.host}:{args.port}"
    log.info("仅测量含 TCP 网络的数据 %s trials=%d", endpoint, args.trials)

    trials_data: list[dict] = []
    all_raw: list[dict] = []
    for i in range(1, args.trials + 1):
        try:
            row = run_one_trial(i, args.password, args.sizes, args.host, args.port)
            trials_data.append(row)
            all_raw.append(row)
        except Exception as e:
            log.error("trial 失败 i=%d: %s", i, e)
            fail = {
                "protocol": "wbp",
                "display": PROTOCOL_DISPLAY,
                "trial_id": i,
                "idc": f"e2e_wbp_{i}",
                "init_ms": 0.0,
                "rec_ms": 0.0,
                "init_comm_bytes": 0,
                "rec_comm_bytes": 0,
                "key_match": False,
                "enc_times": {},
                "dec_times": {},
                "success": False,
                "error": str(e),
            }
            trials_data.append(fail)
            all_raw.append(fail)

    aggregates = [aggregate_results(trials_data, args.sizes)]
    meta = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "url": endpoint,
    }
    path = write_excel(aggregates, all_raw, args.sizes, Path(args.output), meta)
    print_summary(aggregates, path, endpoint)


if __name__ == "__main__":
    main()
