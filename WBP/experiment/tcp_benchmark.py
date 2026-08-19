"""WBP TCP 基线测试（对齐 PAEE ``experiment/tcp_benchmark``）。

本机/远端 TCP，真实 Client↔Server；分阶段统计 **延迟** 与 **通信量**。

对标 PAEE::
  Init ≈ Enc_proto（封装/注册备份密钥）
  Rec  ≈ Dec_proto（恢复备份密钥）
  HELLO/建连在计时外（长连接复用）

轮数：一对 Client↔Server 交互算 1（不含 HELLO）::
  Init=2，Rec=2

用法::

    # 终端1
    python run_server.py --host 0.0.0.0 --port 8765

    # 终端2
    python -m experiment.tcp_benchmark --trials 20 --host 127.0.0.1 --port 8765
    python -m experiment.tcp_benchmark --trials 20 --host 54.x.x.x --port 8765 -q

输出（与 PAEE 同名，便于对照）::
    experiment/output/tcp_network_benchmark.xlsx
    experiment/output/tcp_network_benchmark.md
    experiment/output/tcp_network_benchmark.csv
"""

from __future__ import annotations

import argparse
import csv
import logging
import secrets
import socket
import statistics
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from client.app import DEFAULT_HOST, DEFAULT_PORT, INIT_ROUNDS, REC_ROUNDS, Client

log = logging.getLogger("wbp.bench")

DISPLAY = "WBP (DFG+23)"

# 阶段名带对标标注，便于与 PAEE Enc_proto / Dec_proto 对照
PHASE_SPECS = (
    ("Init≈Enc_proto", "init_latency_ms", "init_comm_bytes", INIT_ROUNDS),
    ("Rec≈Dec_proto", "rec_latency_ms", "rec_comm_bytes", REC_ROUNDS),
)

PROTOCOL_META = {
    "wbp": {
        "display": DISPLAY,
        "rounds_init": INIT_ROUNDS,
        "rounds_rec": REC_ROUNDS,
    },
}


def _mean_std(values: list[float]) -> tuple[float, float]:
    if not values:
        return 0.0, 0.0
    if len(values) == 1:
        return values[0], 0.0
    return statistics.mean(values), statistics.stdev(values)


def check_server(host: str, port: int) -> None:
    """探测 TCP Server 是否可达。"""
    try:
        with socket.create_connection((host, port), timeout=3):
            pass
    except OSError as e:
        raise SystemExit(
            f"无法连接 WBP Server {host}:{port}，请先启动: python run_server.py\n原因: {e}"
        ) from e


def run_trials(
    n: int, password: str, host: str, port: int, idc_prefix: str
) -> list[dict]:
    """跑 n 次完整 Init+Rec（TCP），建连在阶段计时外；返回原始 trial 列表。"""
    rows: list[dict] = []
    meta = PROTOCOL_META["wbp"]
    for i in range(1, n + 1):
        idc = f"{idc_prefix}_{i}"
        log.info("---------- WBP trial %d/%d idc=%s ----------", i, n, idc)
        client: Client | None = None
        try:
            client = Client(idc=idc, host=host, port=port, password=password)
            client.connect()  # HELLO：不计时、不计入 Init/Rec 通信量
            init_r, init_m = client.init_metrics()
            if not init_r.ok:
                raise RuntimeError(f"Init 失败: {init_r.error}")
            rec_r, rec_m = client.recover_metrics()
            if not rec_r.ok:
                raise RuntimeError(f"Rec 失败: {rec_r.error}")
            key_match = init_r.backup_key == rec_r.backup_key
            rows.append(
                {
                    "protocol": "wbp",
                    "display": meta["display"],
                    "trial_id": i,
                    "idc": idc,
                    "success": key_match,
                    "key_match": key_match,
                    "init_latency_ms": init_m["latency_ms"],
                    "rec_latency_ms": rec_m["latency_ms"],
                    "total_latency_ms": init_m["latency_ms"] + rec_m["latency_ms"],
                    "init_comm_bytes": init_m["comm_bytes"],
                    "rec_comm_bytes": rec_m["comm_bytes"],
                    "total_comm_bytes": init_m["comm_bytes"] + rec_m["comm_bytes"],
                    "init_rounds": meta["rounds_init"],
                    "rec_rounds": meta["rounds_rec"],
                    "error": "",
                }
            )
            log.info(
                "OK init=%.2fms/%dB rec=%.2fms/%dB match=%s",
                init_m["latency_ms"],
                init_m["comm_bytes"],
                rec_m["latency_ms"],
                rec_m["comm_bytes"],
                key_match,
            )
        except Exception as e:
            log.error("trial %d 失败: %s", i, e)
            rows.append(
                {
                    "protocol": "wbp",
                    "display": meta["display"],
                    "trial_id": i,
                    "idc": idc,
                    "success": False,
                    "key_match": False,
                    "init_latency_ms": 0.0,
                    "rec_latency_ms": 0.0,
                    "total_latency_ms": 0.0,
                    "init_comm_bytes": 0,
                    "rec_comm_bytes": 0,
                    "total_comm_bytes": 0,
                    "init_rounds": meta["rounds_init"],
                    "rec_rounds": meta["rounds_rec"],
                    "error": str(e),
                }
            )
        finally:
            if client is not None:
                client.close()
    return rows


def aggregate(rows: list[dict]) -> dict[str, Any]:
    """对成功 trial 按阶段求 mean±std（结构对齐 PAEE）。"""
    ok = [r for r in rows if r["success"]]
    phases: dict[str, dict[str, float]] = {}
    for name, lat_k, comm_k, rounds in PHASE_SPECS:
        lm, ls = _mean_std([float(r[lat_k]) for r in ok])
        cm, cs = _mean_std([float(r[comm_k]) for r in ok])
        phases[name] = {
            "rounds": rounds,
            "latency_mean_ms": lm,
            "latency_std_ms": ls,
            "comm_mean_bytes": cm,
            "comm_std_bytes": cs,
        }

    # 兼容旧字段名（password_length_bench / 旧脚本）
    init_p = phases["Init≈Enc_proto"]
    rec_p = phases["Rec≈Dec_proto"]
    tot_lat = [r["total_latency_ms"] for r in ok]
    tot_comm = [float(r["total_comm_bytes"]) for r in ok]
    tot_lm, tot_ls = _mean_std(tot_lat)
    tot_cm, tot_cs = _mean_std(tot_comm)

    return {
        "protocol": "wbp",
        "display": DISPLAY,
        "n_trials": len(rows),
        "n_success": len(ok),
        "success_rate": len(ok) / len(rows) if rows else 0.0,
        "phases": phases,
        "init_rounds": INIT_ROUNDS,
        "rec_rounds": REC_ROUNDS,
        "init_latency_mean_ms": init_p["latency_mean_ms"],
        "init_latency_std_ms": init_p["latency_std_ms"],
        "rec_latency_mean_ms": rec_p["latency_mean_ms"],
        "rec_latency_std_ms": rec_p["latency_std_ms"],
        "total_latency_mean_ms": tot_lm,
        "total_latency_std_ms": tot_ls,
        "init_comm_mean_bytes": init_p["comm_mean_bytes"],
        "init_comm_std_bytes": init_p["comm_std_bytes"],
        "rec_comm_mean_bytes": rec_p["comm_mean_bytes"],
        "rec_comm_std_bytes": rec_p["comm_std_bytes"],
        "total_comm_mean_bytes": tot_cm,
        "total_comm_std_bytes": tot_cs,
    }


NOTE = (
    "口径：长连接复用；建连/HELLO 不计时。"
    "轮数=交互轮(一对C↔S算1，不含HELLO)。"
    "Init≈PAEE Enc_proto；Rec≈PAEE Dec_proto。"
    "通信量=二进制 TCP wire（PBCS风格 opcode+u16LV；无JSON/Base64/外层长度头）。"
)


def write_outputs(
    agg: dict[str, Any],
    raw: list[dict],
    out_dir: Path,
    *,
    endpoint: str,
    timestamp: str,
) -> Path:
    """写 md / csv / xlsx（文件名对齐 PAEE）。"""
    out_dir.mkdir(parents=True, exist_ok=True)

    md = out_dir / "tcp_network_benchmark.md"
    lines = [
        "# WBP TCP 基线（对标 PAEE Enc_proto / Dec_proto）",
        "",
        f"生成时间: {timestamp}",
        f"trials={agg['n_trials']} success={agg['n_success']} "
        f"({agg['success_rate'] * 100:.1f}%)",
        f"Server: `{endpoint}`",
        "",
        NOTE,
        "",
        "| 阶段 | 轮数 | 延迟 mean±std (ms) | 通信量 mean±std (B) |",
        "|------|------|--------------------|---------------------|",
    ]
    for phase, p in agg["phases"].items():
        lines.append(
            f"| {phase} | {p['rounds']} | "
            f"{p['latency_mean_ms']:.2f} ± {p['latency_std_ms']:.2f} | "
            f"{p['comm_mean_bytes']:.1f} ± {p['comm_std_bytes']:.1f} |"
        )
    md.write_text("\n".join(lines) + "\n", encoding="utf-8")

    csv_path = out_dir / "tcp_network_benchmark.csv"
    if raw:
        with csv_path.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(raw[0].keys()))
            w.writeheader()
            w.writerows(raw)

    xlsx_path = out_dir / "tcp_network_benchmark.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"
    headers = [
        "方案",
        "阶段",
        "试验次数",
        "成功次数",
        "成功率(%)",
        "轮数",
        "延迟_mean(ms)",
        "延迟_std(ms)",
        "通信量_mean(bytes)",
        "通信量_std(bytes)",
        "Server 端点",
        "测量说明",
        "时间戳",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for phase, p in agg["phases"].items():
        ws.append(
            [
                agg["display"],
                phase,
                agg["n_trials"],
                agg["n_success"],
                round(agg["success_rate"] * 100, 1),
                p["rounds"],
                round(p["latency_mean_ms"], 3),
                round(p["latency_std_ms"], 3),
                round(p["comm_mean_bytes"], 1),
                round(p["comm_std_bytes"], 1),
                endpoint,
                NOTE,
                timestamp,
            ]
        )
    # Total 行（PAEE 基线无此行；保留便于总览）
    ws.append(
        [
            agg["display"],
            "Total",
            agg["n_trials"],
            agg["n_success"],
            round(agg["success_rate"] * 100, 1),
            agg["init_rounds"] + agg["rec_rounds"],
            round(agg["total_latency_mean_ms"], 3),
            round(agg["total_latency_std_ms"], 3),
            round(agg["total_comm_mean_bytes"], 1),
            round(agg["total_comm_std_bytes"], 1),
            endpoint,
            NOTE,
            timestamp,
        ]
    )

    ws2 = wb.create_sheet("原始数据")
    raw_headers = [
        "方案",
        "trial_id",
        "idc",
        "success",
        "Init延迟(ms)",
        "Rec延迟(ms)",
        "Init通信(B)",
        "Rec通信(B)",
        "error",
    ]
    ws2.append(raw_headers)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for r in raw:
        ws2.append(
            [
                r["display"],
                r["trial_id"],
                r["idc"],
                r["success"],
                round(r["init_latency_ms"], 3),
                round(r["rec_latency_ms"], 3),
                r["init_comm_bytes"],
                r["rec_comm_bytes"],
                r.get("error", ""),
            ]
        )

    for sheet in (ws, ws2):
        for col in sheet.columns:
            letter = col[0].column_letter
            width = min(44, max(12, max(len(str(c.value or "")) for c in col) + 2))
            sheet.column_dimensions[letter].width = width

    wb.save(xlsx_path)
    log.info("wrote outputs under %s", out_dir)
    return xlsx_path


def print_summary(agg: dict[str, Any], out_dir: Path) -> None:
    print(f"\n=== {DISPLAY} protocol baseline (vs PAEE Enc/Dec_proto) ===")
    print(
        f"trials={agg['n_trials']} success={agg['n_success']} "
        f"({agg['success_rate'] * 100:.1f}%)"
    )
    for phase, p in agg["phases"].items():
        print(
            f"  {phase:18s}  "
            f"{p['latency_mean_ms']:8.2f}±{p['latency_std_ms']:6.2f} ms  "
            f"{p['comm_mean_bytes']:8.1f}±{p['comm_std_bytes']:6.1f} B  "
            f"rounds={p['rounds']}"
        )
    print(f"\n结果目录: {out_dir}")
    print("  tcp_network_benchmark.xlsx / .md / .csv\n")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="WBP TCP 基线：分阶段延迟+通信量（对齐 PAEE tcp_benchmark）"
    )
    parser.add_argument("--trials", type=int, default=20, help="试验次数（默认20）")
    parser.add_argument("--password", default="benchmark_pw_2024")
    parser.add_argument("--host", default=DEFAULT_HOST)
    parser.add_argument("--port", type=int, default=DEFAULT_PORT)
    parser.add_argument("--idc-prefix", default="tcp_wbp")
    parser.add_argument(
        "--out-dir",
        type=str,
        default=str(ROOT / "experiment" / "output"),
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="兼容旧参数：xlsx 路径；默认写到 --out-dir/tcp_network_benchmark.xlsx",
    )
    parser.add_argument("-q", "--quiet", action="store_true", help="减少 wire 日志")
    args = parser.parse_args()

    level = logging.WARNING if args.quiet else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s %(name)s %(levelname)s %(message)s",
    )
    if args.quiet:
        for name in ("wbp.wire", "wbp.client", "wbp.server", "wbp.hsm"):
            logging.getLogger(name).setLevel(logging.WARNING)

    host = args.host
    if host in ("0.0.0.0", "::"):
        host = "127.0.0.1"

    check_server(host, args.port)
    endpoint = f"{host}:{args.port}"
    run_tag = datetime.now().strftime("%Y%m%d%H%M%S") + secrets.token_hex(2)
    idc_prefix = f"{args.idc_prefix}_{run_tag}"
    log.info("开始 WBP TCP 基线 trials=%d %s idc_prefix=%s", args.trials, endpoint, idc_prefix)

    raw = run_trials(args.trials, args.password, host, args.port, idc_prefix=idc_prefix)
    agg = aggregate(raw)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    out_dir = Path(args.out_dir)
    if args.output:
        # 旧 --output 指定 xlsx 时，out_dir 取其所在目录
        out_dir = Path(args.output).resolve().parent
    write_outputs(agg, raw, out_dir, endpoint=endpoint, timestamp=ts)
    print_summary(agg, out_dir)
    raise SystemExit(0 if agg["n_success"] == agg["n_trials"] else 1)


if __name__ == "__main__":
    main()
