# -*- coding: utf-8 -*-
"""
experiment/file_size_proto_bench.py
===================================
【实验三】文件大小扫描（含网络）：口令长度固定 16；大小 1…500 MB。

口径：**不传 π**；Enc = Ext+Wrap+AES-CTR→ct2+H5(...,ct2)+ENC_COMMIT（ct2 不上云）；
Dec = Ext + 取 (ct0,ct1,τ,d) + 本地 ct2 验 τ + AES-CTR.Dec。

用法::
    python -m experiment.file_size_proto_bench --trials 3 --host 35.78.207.231 --port 5202
"""

from __future__ import annotations

import argparse
import csv
import logging
import os
import secrets
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from experiment.tcp_benchmark_lib import (  # noqa: E402
    check_server,
    mean_std,
    start_auto_server,
)

log = logging.getLogger("paee.filesize")
DISPLAY = "PAEE (Fig.1)"
DEFAULT_SIZES_MB = (1, 10, 100, 200, 300, 400, 500)
PASSWORD_LEN = 16


def make_password(length: int = PASSWORD_LEN) -> str:
    return "a" * length


def _gen(size_bytes: int) -> bytes:
    chunk = 8 * 1024 * 1024
    parts: List[bytes] = []
    left = size_bytes
    while left > 0:
        n = min(chunk, left)
        parts.append(os.urandom(n))
        left -= n
    return b"".join(parts)


def run_size_trials(
    *,
    host: str,
    port: int,
    password: str,
    size_mb: int,
    trials: int,
    lambda_bytes: int,
    id_prefix: str,
    save_ct_dir: Optional[Path],
    tau_bind_ct2: bool = True,
) -> List[Dict[str, Any]]:
    """
    每个 trial：
      Reg → Ext → Enc → Dec（ct2 本地不上云）。
    tau_bind_ct2：True → τ=H5(kMAC,(ct0,ct1,ct2))；False → τ=H5(kMAC,(ct0,ct1))。
    """
    from net.client_api import PAEEClientSession
    from net.client_metrics import (
        decrypt_compare_metrics,
        encrypt_compare_metrics,
        register_metrics,
    )
    from paee.params import Setup

    pp = Setup(lambda_bytes)
    size = size_mb * 1024 * 1024
    log.info("生成明文 %d MB ...", size_mb)
    plaintext = _gen(size)
    rows: List[Dict[str, Any]] = []
    h5_tag = "H5(ct0,ct1,ct2)" if tau_bind_ct2 else "H5(ct0,ct1)"

    for t in range(1, trials + 1):
        uid = f"{id_prefix}_{size_mb}mb_{t}"
        log.info("---- %d MB trial %d/%d id=%s ----", size_mb, t, trials, uid)
        try:
            with PAEEClientSession(host, port, pp) as sess:
                pk_out, _reg = register_metrics(
                    host, port, pp, uid, password, session=sess
                )
                enc = encrypt_compare_metrics(
                    host,
                    port,
                    pp,
                    pk_out,
                    uid,
                    password,
                    plaintext=plaintext,
                    session=sess,
                    tau_bind_ct2=tau_bind_ct2,
                )
                ext_ms = float(enc["ext"]["latency_ms"])
                enc_total_ms = float(enc["enc_full"]["latency_ms"])
                enc_file_ms = float(enc["commit"]["encm_ms"])
                # Wrap+H5+网络（不含纯 Encm）
                enc_key_ms = float(enc["commit"]["proto_ms"])
                local_ct2 = enc["ct"].ct2

                if save_ct_dir is not None:
                    save_ct_dir.mkdir(parents=True, exist_ok=True)
                    (save_ct_dir / f"{uid}.ct2").write_bytes(local_ct2)

                dec = decrypt_compare_metrics(
                    host,
                    port,
                    pp,
                    pk_out,
                    uid,
                    password,
                    local_ct2=local_ct2,
                    session=sess,
                    tau_bind_ct2=tau_bind_ct2,
                )
            if dec["plaintext"] != plaintext:
                raise RuntimeError("plaintext mismatch")

            dec_total_ms = float(dec["dec_full"]["latency_ms"])
            dec_proto_ms = float(dec["dec_proto"]["latency_ms"])
            dec_file_ms = float(dec["dec_full"]["dec_m_ms"])

            row = {
                "success": True,
                "size_mb": size_mb,
                "trial_id": t,
                "id": uid,
                "ext_ms": ext_ms,
                "enc_key_store_ms": enc_key_ms,
                "enc_file_ms": enc_file_ms,
                "enc_total_ms": enc_total_ms,
                "dec_proto_ms": dec_proto_ms,
                "dec_file_ms": dec_file_ms,
                "dec_total_ms": dec_total_ms,
                "enc_comm_bytes": int(enc["enc_full"]["comm_bytes"]),
                "dec_comm_bytes": int(dec["dec_full"]["comm_bytes"]),
                "ct_bytes": len(local_ct2),
                "error": "",
            }
            rows.append(row)
            log.info(
                "  Ext=%.1f Enc=%.1f(%s) Dec=%.1f ct2=%dB wire_enc=%dB",
                ext_ms,
                enc_total_ms,
                h5_tag,
                dec_total_ms,
                len(local_ct2),
                int(enc["enc_full"]["comm_bytes"]),
            )
            del local_ct2
        except Exception as e:  # noqa: BLE001
            log.error("trial fail: %s", e)
            rows.append(
                {
                    "success": False,
                    "size_mb": size_mb,
                    "trial_id": t,
                    "id": uid,
                    "ext_ms": 0.0,
                    "enc_key_store_ms": 0.0,
                    "enc_file_ms": 0.0,
                    "enc_total_ms": 0.0,
                    "dec_proto_ms": 0.0,
                    "dec_file_ms": 0.0,
                    "dec_total_ms": 0.0,
                    "enc_comm_bytes": 0,
                    "dec_comm_bytes": 0,
                    "ct_bytes": 0,
                    "error": str(e),
                }
            )

    del plaintext
    return rows


def phase_rows(tau_bind_ct2: bool = True):
    if tau_bind_ct2:
        return (
            ("Ext", "ext_ms"),
            ("Enc（Ext+封装+H5(ct2)+加密文件；ct2不上云）", "enc_total_ms"),
            ("Dec（Ext+取材料+验τ(ct2)+解密；ct2本地）", "dec_total_ms"),
        )
    return (
        ("Ext", "ext_ms"),
        ("Enc（Ext+封装+H5(ct0,ct1)+加密文件；ct2不上云）", "enc_total_ms"),
        ("Dec（Ext+取材料+验τ(ct0,ct1)+解密；ct2本地）", "dec_total_ms"),
    )


def breakdown_rows(tau_bind_ct2: bool = True):
    h5 = "H5(ct2)" if tau_bind_ct2 else "H5(ct0,ct1)"
    return (
        ("Ext", "ext_ms"),
        (f"Enc_key（Wrap+{h5}+COMMIT，proto_ms）", "enc_key_store_ms"),
        ("Enc_file（AES-CTR→ct2）", "enc_file_ms"),
        ("Enc_total", "enc_total_ms"),
        ("Dec_proto（Ext+至dek，含验τ）", "dec_proto_ms"),
        ("Dec_file（SE.Dec ct2）", "dec_file_ms"),
        ("Dec_total", "dec_total_ms"),
    )


PHASE_ROWS = phase_rows(True)
BREAKDOWN_ROWS = breakdown_rows(True)


def summarize_by_size(
    all_raw: List[Dict[str, Any]],
    sizes: Sequence[int],
    *,
    tau_bind_ct2: bool = True,
) -> Dict[int, Dict[str, Any]]:
    rows_spec = breakdown_rows(tau_bind_ct2)
    out: Dict[int, Dict[str, Any]] = {}
    for mb in sizes:
        rows = [r for r in all_raw if r["size_mb"] == mb]
        ok = [r for r in rows if r["success"]]
        stats: Dict[str, Any] = {
            "n_trials": len(rows),
            "n_success": len(ok),
        }
        for _name, key in rows_spec:
            m, s = mean_std([float(r[key]) for r in ok])
            stats[f"{key}_mean"] = m
            stats[f"{key}_std"] = s
        out[mb] = stats
    return out


def _size_header(sizes: Sequence[int]) -> List[str]:
    return ["阶段"] + [f"{mb}MB" for mb in sizes]


def _row_mean_std(
    by_size: Dict[int, Dict[str, Any]],
    sizes: Sequence[int],
    phase: str,
    key: str,
) -> List[Any]:
    cells: List[Any] = [phase]
    for mb in sizes:
        st = by_size[mb]
        cells.append(
            f"{st[f'{key}_mean']:.2f} ± {st[f'{key}_std']:.2f}"
        )
    return cells


def write_outputs(
    by_size: Dict[int, Dict[str, Any]],
    raw: List[Dict[str, Any]],
    sizes: Sequence[int],
    out_dir: Path,
    *,
    endpoint: str,
    trials: int,
    password_len: int,
    timestamp: str,
    tau_bind_ct2: bool = True,
    out_stem: str = "file_size_proto_benchmark",
) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    phases = phase_rows(tau_bind_ct2)
    brk = breakdown_rows(tau_bind_ct2)
    if tau_bind_ct2:
        note = (
            f"【实验三】口令长度={password_len}；trials/大小={trials}；Server={endpoint}；"
            "**不传π**；τ=H5(kMAC,(ct0,ct1,ct2))；ct2仅本地不上云；"
            "Enc=Ext+Wrap+Encm+H5+COMMIT；Dec=Ext+取材料+验τ+Dec ct2"
        )
        title = "# PAEE 文件大小：Ext / Enc / Dec 延迟"
    else:
        note = (
            f"【实验三·副本】口令长度={password_len}；trials/大小={trials}；"
            f"Server={endpoint}；**不传π**；"
            "**仅改** τ=H5(kMAC,(ct0,ct1))（不含 ct2）；ct2仍本地不上云；"
            "其余同实验三"
        )
        title = "# PAEE 文件大小：Ext / Enc / Dec（τ 不含 ct2）"
    header = _size_header(sizes)
    sep = "|" + "|".join(["------"] * len(header)) + "|"

    md = [
        title,
        "",
        f"生成时间: {timestamp}",
        note,
        "",
        "## 主表 mean±std (ms)",
        "",
        "| " + " | ".join(header) + " |",
        sep,
    ]
    for phase, key in phases:
        md.append(
            "| "
            + " | ".join(
                str(c) for c in _row_mean_std(by_size, sizes, phase, key)
            )
            + " |"
        )
    md += ["", "## 分解 mean±std (ms)", "", "| " + " | ".join(header) + " |", sep]
    for phase, key in brk:
        md.append(
            "| "
            + " | ".join(
                str(c) for c in _row_mean_std(by_size, sizes, phase, key)
            )
            + " |"
        )
    (out_dir / f"{out_stem}.md").write_text(
        "\n".join(md) + "\n", encoding="utf-8"
    )

    with (out_dir / f"{out_stem}_summary.csv").open(
        "w", newline="", encoding="utf-8"
    ) as f:
        w = csv.writer(f)
        w.writerow(header)
        for phase, key in phases:
            w.writerow(_row_mean_std(by_size, sizes, phase, key))
        w.writerow([])
        w.writerow(["分解"])
        for phase, key in brk:
            w.writerow(_row_mean_std(by_size, sizes, phase, key))

    if raw:
        with (out_dir / f"{out_stem}.csv").open(
            "w", newline="", encoding="utf-8"
        ) as f:
            w = csv.DictWriter(f, fieldnames=list(raw[0].keys()))
            w.writeheader()
            w.writerows(raw)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        log.warning("openpyxl missing; skip xlsx")
        return

    def style_h(ws) -> None:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"
    ws.append(header)
    style_h(ws)
    for phase, key in phases:
        ws.append(_row_mean_std(by_size, sizes, phase, key))
    ws.append([])
    ws.append([note])
    ws.append([f"方案={DISPLAY}", f"时间={timestamp}"])

    ws2 = wb.create_sheet("分解")
    ws2.append(header)
    style_h(ws2)
    for phase, key in brk:
        ws2.append(_row_mean_std(by_size, sizes, phase, key))

    ws3 = wb.create_sheet("延迟_mean")
    ws3.append(header)
    style_h(ws3)
    for phase, key in phases:
        ws3.append(
            [phase] + [round(by_size[mb][f"{key}_mean"], 3) for mb in sizes]
        )

    ws4 = wb.create_sheet("原始数据")
    if raw:
        keys = list(raw[0].keys())
        ws4.append(keys)
        style_h(ws4)
        for r in raw:
            ws4.append([r[k] for k in keys])

    path = out_dir / f"{out_stem}.xlsx"
    wb.save(path)
    log.info("wrote %s", path)


def main(
    argv: Optional[List[str]] = None,
    *,
    tau_bind_ct2: bool = True,
    out_stem: str = "file_size_proto_benchmark",
    id_prefix_base: str = "fsize",
    description: str = "PAEE Ext / Enc / Dec vs file size (pw len=16)",
) -> int:
    parser = argparse.ArgumentParser(description=description)
    parser.add_argument("--trials", type=int, default=3)
    parser.add_argument(
        "--sizes",
        type=int,
        nargs="+",
        default=list(DEFAULT_SIZES_MB),
        help="文件大小 MB，默认 1 10 100 200 300 400 500",
    )
    parser.add_argument("--password-len", type=int, default=PASSWORD_LEN)
    parser.add_argument("--host", default=None)
    parser.add_argument("--port", type=int, default=None)
    parser.add_argument("--auto-server", action="store_true")
    parser.add_argument(
        "--save-ct",
        action="store_true",
        help="将文件密文写到 experiment/output/file_cts/（不计时）",
    )
    parser.add_argument(
        "--out-dir", default=str(ROOT / "experiment" / "output")
    )
    parser.add_argument("-q", "--quiet", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )

    from main import load_config

    cfg = load_config()
    lambda_bytes = int(cfg["crypto"]["lambda_bytes"])
    password = make_password(args.password_len)
    assert len(password) == args.password_len

    host = args.host if args.host is not None else str(cfg["network"]["host"])
    port = args.port if args.port is not None else int(cfg["network"]["port"])
    if host in ("0.0.0.0", "::"):
        host = "127.0.0.1"

    stop = None
    if args.auto_server:
        host, port, stop = start_auto_server(
            ROOT / "data" / "bench_filesize_server"
        )
    else:
        check_server(host, port)

    run_tag = datetime.now().strftime("%Y%m%d%H%M%S") + secrets.token_hex(2)
    id_prefix = f"{id_prefix_base}_{run_tag}"
    save_ct_dir = (
        Path(args.out_dir) / "file_cts" / run_tag if args.save_ct else None
    )
    phases = phase_rows(tau_bind_ct2)

    all_raw: List[Dict[str, Any]] = []
    try:
        for mb in args.sizes:
            log.info("===== file size %d MB =====", mb)
            all_raw.extend(
                run_size_trials(
                    host=host,
                    port=port,
                    password=password,
                    size_mb=mb,
                    trials=args.trials,
                    lambda_bytes=lambda_bytes,
                    id_prefix=id_prefix,
                    save_ct_dir=save_ct_dir,
                    tau_bind_ct2=tau_bind_ct2,
                )
            )

        by_size = summarize_by_size(
            all_raw, args.sizes, tau_bind_ct2=tau_bind_ct2
        )
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        write_outputs(
            by_size,
            all_raw,
            args.sizes,
            Path(args.out_dir),
            endpoint=f"{host}:{port}",
            trials=args.trials,
            password_len=args.password_len,
            timestamp=ts,
            tau_bind_ct2=tau_bind_ct2,
            out_stem=out_stem,
        )

        header = _size_header(args.sizes)
        col_w = max(16, max(len(h) for h in header[1:]) + 2)
        print()
        print(f"{header[0]:<48}" + "".join(f"{h:>{col_w}}" for h in header[1:]))
        for phase, key in phases:
            row = _row_mean_std(by_size, args.sizes, phase, key)
            print(
                f"{row[0]:<48}" + "".join(f"{c:>{col_w}}" for c in row[1:])
            )
        print(f"\n结果: {args.out_dir}/{out_stem}.xlsx|.md")
    finally:
        if stop is not None:
            stop.set()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
