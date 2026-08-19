# -*- coding: utf-8 -*-
"""
experiment/file_size_crypto_bench.py
====================================
【实验四】文件大小扫描（纯密码学、无网络）：口令 16；大小 1…500 MB。

口径：**不传/不算 π**（本地 eval 亦不 Prove）；
  Ext = 盲化 + POPRF 求值 + finalize
  Enc = Ext + Wrap + AES-CTR→ct2 + H5(...,ct2) + 本地 store
  Dec = Ext + SDec + 本地 ct2 验 τ + AES-CTR.Dec

用法::
    python -m experiment.file_size_crypto_bench --trials 5
"""

from __future__ import annotations

import argparse
import csv
import logging
import os
import secrets
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from experiment.tcp_benchmark_lib import mean_std  # noqa: E402

log = logging.getLogger("paee.crypto-filesize")
DISPLAY = "PAEE (Fig.1) crypto-only"
DEFAULT_SIZES_MB = (1, 10, 100, 200, 300, 400, 500)
PASSWORD_LEN = 16


def make_password(length: int = PASSWORD_LEN) -> str:
    return "a" * length


def _gen(size_bytes: int) -> bytes:
    chunk = 8 * 1024 * 1024
    parts: List[bytes] = []
    left = size_bytes
    while left > 0:
        n = min(chunk, left)
        parts.append(os.urandom(n))
        left -= n
    return b"".join(parts)


def _register(state, pp, uid: str, pw: str, pk) -> None:
    """本地 Reg（不计目标表）。"""
    from paee import protocol as proto

    ctx = state.SReg_issue_ctx(uid)
    if ctx is None:
        raise RuntimeError(f"Reg issue failed for {uid}")
    st = proto.CReg_blind(pw)
    ev = state.SReg_eval(uid, st.a)
    if ev is None:
        raise RuntimeError("Reg eval failed")
    a_tilde = ev
    c = proto.CReg_finalize_c(pp, uid, pw, ctx, st, a_tilde, pk.K)
    if c is None or not state.SReg_store(uid, c):
        raise RuntimeError("Reg store failed")


def _ext_crypto(state, pp, uid: str, pw: str, pk):
    """
    Ext 全链路密码学（无网络）：
      Client blind → Server eval(ã) → Client finalize（无 π）。
    """
    from paee import protocol as proto

    ctx = state.SExt_ctx(uid)
    if ctx is None:
        raise RuntimeError("Ext: unknown id")
    st = proto.CReg_blind(pw)
    a_tilde = state.SExt_eval(uid, st.a)
    if a_tilde is None:
        raise RuntimeError("Ext eval failed")
    est = proto.CExt(pp, uid, pw, ctx, st, a_tilde, pk.K)
    if est is None:
        raise RuntimeError("Ext finalize failed")
    return est


def run_size_trials(
    *,
    password: str,
    size_mb: int,
    trials: int,
    lambda_bytes: int,
    id_prefix: str,
    tau_bind_ct2: bool = True,
) -> List[Dict[str, Any]]:
    from paee import envelope
    from paee.params import SerKGen, Setup
    from paee.protocol import PAEEServerState

    pp = Setup(lambda_bytes)
    size = size_mb * 1024 * 1024
    log.info("生成明文 %d MB ...", size_mb)
    plaintext = _gen(size)
    rows: List[Dict[str, Any]] = []
    h5_tag = "H5(ct0,ct1,ct2)" if tau_bind_ct2 else "H5(ct0,ct1)"

    for t in range(1, trials + 1):
        uid = f"{id_prefix}_{size_mb}mb_{t}"
        log.info("---- %d MB trial %d/%d id=%s ----", size_mb, t, trials, uid)
        try:
            sk = SerKGen(pp)
            pk = sk  # ServerKey 含 K,X
            state = PAEEServerState(pp, sk)
            _register(state, pp, uid, password, pk)

            # 1) Ext only（纯计算）
            t0 = time.perf_counter()
            _ = _ext_crypto(state, pp, uid, password, pk)
            ext_ms = (time.perf_counter() - t0) * 1000.0

            # 2) Enc = Ext + Wrap + Encm + H5 + 本地 store（无网络）
            t1 = time.perf_counter()
            est_enc = _ext_crypto(state, pp, uid, password, pk)
            out = envelope.Enc(
                pp,
                pk,
                uid,
                password,
                est_enc,
                plaintext,
                tau_bind_ct2=tau_bind_ct2,
            )
            if out is None:
                raise RuntimeError("Enc failed")
            c_prime, ct = out
            from paee.types import Ciphertext

            ct_server = Ciphertext(
                ct0=ct.ct0, ct1=ct.ct1, ct2=b"", tau=ct.tau
            )
            if not state.Enc_store(uid, c_prime, ct_server):
                raise RuntimeError("Enc_store failed")
            local_ct2 = ct.ct2
            enc_ms = (time.perf_counter() - t1) * 1000.0

            # 3) Dec = Ext + SDec + 本地 ct2 合并验 τ + 解密
            t2 = time.perf_counter()
            est_dec = _ext_crypto(state, pp, uid, password, pk)
            resp = state.SDec(uid)
            if resp is None:
                raise RuntimeError("SDec failed")
            ct_s, d = resp
            ct_full = Ciphertext(
                ct0=ct_s.ct0, ct1=ct_s.ct1, ct2=local_ct2, tau=ct_s.tau
            )
            m = envelope.client_dec(
                pp,
                password,
                est_dec,
                ct_full,
                d,
                tau_bind_ct2=tau_bind_ct2,
            )
            if m != plaintext:
                raise RuntimeError("plaintext mismatch")
            dec_ms = (time.perf_counter() - t2) * 1000.0

            rows.append(
                {
                    "success": True,
                    "size_mb": size_mb,
                    "trial_id": t,
                    "id": uid,
                    "ext_ms": ext_ms,
                    "enc_ms": enc_ms,
                    "dec_ms": dec_ms,
                    "ct_bytes": len(local_ct2),
                    "error": "",
                }
            )
            log.info(
                "  Ext=%.2fms Enc=%.2fms Dec=%.2fms |ct2|=%d (%s)",
                ext_ms,
                enc_ms,
                dec_ms,
                len(local_ct2),
                h5_tag,
            )
            del ct, ct_server, ct_full, local_ct2, m, out
        except Exception as e:  # noqa: BLE001
            log.error("trial fail: %s", e)
            rows.append(
                {
                    "success": False,
                    "size_mb": size_mb,
                    "trial_id": t,
                    "id": uid,
                    "ext_ms": 0.0,
                    "enc_ms": 0.0,
                    "dec_ms": 0.0,
                    "ct_bytes": 0,
                    "error": str(e),
                }
            )

    del plaintext
    return rows


def phase_rows(tau_bind_ct2: bool = True):
    if tau_bind_ct2:
        return (
            ("Ext", "ext_ms"),
            ("Enc（Ext+封装+H5(ct2)+加密文件）", "enc_ms"),
            ("Dec（Ext+恢复+验τ(ct2)+解密）", "dec_ms"),
        )
    return (
        ("Ext", "ext_ms"),
        ("Enc（Ext+封装+H5(ct0,ct1)+加密文件）", "enc_ms"),
        ("Dec（Ext+恢复+验τ(ct0,ct1)+解密）", "dec_ms"),
    )


PHASE_ROWS = phase_rows(True)


def summarize_by_size(
    all_raw: List[Dict[str, Any]],
    sizes: Sequence[int],
    *,
    tau_bind_ct2: bool = True,
) -> Dict[int, Dict[str, Any]]:
    phases = phase_rows(tau_bind_ct2)
    out: Dict[int, Dict[str, Any]] = {}
    for mb in sizes:
        rows = [r for r in all_raw if r["size_mb"] == mb]
        ok = [r for r in rows if r["success"]]
        stats: Dict[str, Any] = {
            "n_trials": len(rows),
            "n_success": len(ok),
        }
        for _name, key in phases:
            m, s = mean_std([float(r[key]) for r in ok])
            stats[f"{key}_mean"] = m
            stats[f"{key}_std"] = s
        out[mb] = stats
    return out


def _size_header(sizes: Sequence[int]) -> List[str]:
    return ["阶段"] + [f"{mb}MB" for mb in sizes]


def _row_mean_std(
    by_size: Dict[int, Dict[str, Any]],
    sizes: Sequence[int],
    phase: str,
    key: str,
) -> List[Any]:
    cells: List[Any] = [phase]
    for mb in sizes:
        st = by_size[mb]
        cells.append(f"{st[f'{key}_mean']:.2f} ± {st[f'{key}_std']:.2f}")
    return cells


def write_outputs(
    by_size: Dict[int, Dict[str, Any]],
    raw: List[Dict[str, Any]],
    sizes: Sequence[int],
    out_dir: Path,
    *,
    trials: int,
    password_len: int,
    timestamp: str,
    tau_bind_ct2: bool = True,
    out_stem: str = "file_size_crypto_benchmark",
) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    phases = phase_rows(tau_bind_ct2)
    if tau_bind_ct2:
        note = (
            f"【实验四】纯密码学（无网络）；口令长度={password_len}；"
            f"trials/大小={trials}；**不传/不算π**；"
            "τ=H5(kMAC,(ct0,ct1,ct2))；ct2仅本地；"
            "Enc=Ext+Wrap+Encm+H5；Dec=Ext+SDec+验τ+Dec；列=文件大小"
        )
        title = "# PAEE 文件大小：Ext / Enc / Dec（纯密码学，无网络）"
    else:
        note = (
            f"【实验四·副本】纯密码学（无网络）；口令长度={password_len}；"
            f"trials/大小={trials}；**不传/不算π**；"
            "**仅改** τ=H5(kMAC,(ct0,ct1))（不含 ct2）；其余同实验四"
        )
        title = "# PAEE 文件大小：Ext / Enc / Dec（纯密码学；τ 不含 ct2）"
    header = _size_header(sizes)
    sep = "|" + "|".join(["------"] * len(header)) + "|"

    md = [
        title,
        "",
        f"生成时间: {timestamp}",
        note,
        "",
        "## 延迟 mean±std (ms)",
        "",
        "| " + " | ".join(header) + " |",
        sep,
    ]
    for phase, key in phases:
        md.append(
            "| "
            + " | ".join(str(c) for c in _row_mean_std(by_size, sizes, phase, key))
            + " |"
        )
    (out_dir / f"{out_stem}.md").write_text(
        "\n".join(md) + "\n", encoding="utf-8"
    )

    with (out_dir / f"{out_stem}_summary.csv").open(
        "w", newline="", encoding="utf-8"
    ) as f:
        w = csv.writer(f)
        w.writerow(header)
        for phase, key in phases:
            w.writerow(_row_mean_std(by_size, sizes, phase, key))

    if raw:
        with (out_dir / f"{out_stem}.csv").open(
            "w", newline="", encoding="utf-8"
        ) as f:
            w = csv.DictWriter(f, fieldnames=list(raw[0].keys()))
            w.writeheader()
            w.writerows(raw)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        log.warning("openpyxl missing; skip xlsx")
        return

    def style_h(ws) -> None:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"
    ws.append(header)
    style_h(ws)
    for phase, key in phases:
        ws.append(_row_mean_std(by_size, sizes, phase, key))
    ws.append([])
    ws.append([note])
    ws.append([f"方案={DISPLAY}", f"时间={timestamp}"])

    ws2 = wb.create_sheet("延迟_mean")
    ws2.append(header)
    style_h(ws2)
    for phase, key in phases:
        ws2.append(
            [phase] + [round(by_size[mb][f"{key}_mean"], 3) for mb in sizes]
        )

    ws3 = wb.create_sheet("原始数据")
    if raw:
        keys = list(raw[0].keys())
        ws3.append(keys)
        style_h(ws3)
        for r in raw:
            ws3.append([r[k] for k in keys])

    path = out_dir / f"{out_stem}.xlsx"
    wb.save(path)
    log.info("wrote %s", path)


def main(
    argv: Optional[List[str]] = None,
    *,
    tau_bind_ct2: bool = True,
    out_stem: str = "file_size_crypto_benchmark",
    id_prefix_base: str = "cryptof",
    description: str = "PAEE Ext/Enc/Dec vs file size (crypto only, no network)",
) -> int:
    parser = argparse.ArgumentParser(description=description)
    parser.add_argument("--trials", type=int, default=5)
    parser.add_argument(
        "--sizes",
        type=int,
        nargs="+",
        default=list(DEFAULT_SIZES_MB),
    )
    parser.add_argument("--password-len", type=int, default=PASSWORD_LEN)
    parser.add_argument(
        "--out-dir", default=str(ROOT / "experiment" / "output")
    )
    parser.add_argument("-q", "--quiet", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )

    from main import load_config

    cfg = load_config()
    lambda_bytes = int(cfg["crypto"]["lambda_bytes"])
    password = make_password(args.password_len)

    run_tag = datetime.now().strftime("%Y%m%d%H%M%S") + secrets.token_hex(2)
    id_prefix = f"{id_prefix_base}_{run_tag}"
    phases = phase_rows(tau_bind_ct2)

    all_raw: List[Dict[str, Any]] = []
    for mb in args.sizes:
        log.info("===== file size %d MB (crypto-only) =====", mb)
        all_raw.extend(
            run_size_trials(
                password=password,
                size_mb=mb,
                trials=args.trials,
                lambda_bytes=lambda_bytes,
                id_prefix=id_prefix,
                tau_bind_ct2=tau_bind_ct2,
            )
        )

    by_size = summarize_by_size(
        all_raw, args.sizes, tau_bind_ct2=tau_bind_ct2
    )
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    write_outputs(
        by_size,
        all_raw,
        args.sizes,
        Path(args.out_dir),
        trials=args.trials,
        password_len=args.password_len,
        timestamp=ts,
        tau_bind_ct2=tau_bind_ct2,
        out_stem=out_stem,
    )

    header = _size_header(args.sizes)
    col_w = max(16, max(len(h) for h in header[1:]) + 2)
    print()
    print(f"{header[0]:<44}" + "".join(f"{h:>{col_w}}" for h in header[1:]))
    for phase, key in phases:
        row = _row_mean_std(by_size, args.sizes, phase, key)
        print(f"{row[0]:<44}" + "".join(f"{c:>{col_w}}" for c in row[1:]))
    print(f"\n结果: {args.out_dir}/{out_stem}.xlsx|.md")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
