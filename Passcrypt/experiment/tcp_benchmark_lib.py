# -*- coding: utf-8 -*-
"""
experiment/tcp_benchmark_lib.py
===============================
PAEE TCP 基准共用工具：统计、探活、自动起 Server、写 Excel。
"""

from __future__ import annotations

import logging
import shutil
import socket
import statistics
import threading
import time
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

log = logging.getLogger("paee.bench.lib")


def mean_std(values: List[float]) -> Tuple[float, float]:
    """样本均值与样本标准差（n=1 时 std=0）。"""
    if not values:
        return 0.0, 0.0
    if len(values) == 1:
        return values[0], 0.0
    return statistics.mean(values), statistics.stdev(values)


def load_oob_pk(path: Path):
    """从带外 JSON 加载 pk=(K,X)。"""
    import json
    from paee import serde

    path = Path(path)
    if not path.is_file():
        raise SystemExit(
            f"missing OOB server pk: {path.resolve()}\n"
            "请从 EC2 拷贝 Server 上的 keys/server_pk.json，例如：\n"
            "  scp <user>@35.78.207.231:/path/to/PassCrypt.../data/server/keys/server_pk.json "
            "./data/client/server_pk.json\n"
            "（需先在 EC2 更新代码并重启 server，才会生成该文件）\n"
            "或：python main.py export-pk --out ./data/client/server_pk.json"
            "（仅当本机已有对应 Server 的 keys/）"
        )
    return serde.import_pk(json.loads(path.read_text(encoding="utf-8")))


def resolve_oob_pk(
    *,
    pk_path: Optional[str],
    auto_data_root: Optional[Path],
    cfg: Optional[dict] = None,
):
    """
    解析带外公钥路径：
      --pk > auto-server data_root/keys/server_pk.json > config paths.server_pk
      > storage.local_root/keys/server_pk.json
    """
    if pk_path:
        return load_oob_pk(Path(pk_path))
    if auto_data_root is not None:
        p = auto_data_root / "keys" / "server_pk.json"
        deadline = time.time() + 5
        while time.time() < deadline:
            if p.is_file():
                return load_oob_pk(p)
            time.sleep(0.05)
        raise SystemExit(f"auto-server pk not ready: {p}")
    if cfg is not None:
        cand = Path(cfg.get("paths", {}).get("server_pk", "data/client/server_pk.json"))
        if cand.is_file():
            return load_oob_pk(cand)
        alt = Path(cfg["storage"]["local_root"]) / "keys" / "server_pk.json"
        if alt.is_file():
            return load_oob_pk(alt)
        raise SystemExit(
            f"missing OOB pk ({cand} or {alt}). "
            "Copy Server keys/server_pk.json or pass --pk."
        )
    raise SystemExit("cannot resolve OOB pk")


def check_server(host: str, port: int) -> None:
    """探测 Server 是否在听；失败则提示先启动或改用 --auto-server。"""
    try:
        with socket.create_connection((host, port), timeout=3):
            pass
    except OSError as e:
        raise SystemExit(
            f"无法连接 PAEE Server {host}:{port}。"
            f"请先 `python main.py server` 或加 `--auto-server`。\n原因: {e}"
        ) from e


def free_port() -> int:
    """向 OS 要一个空闲本机端口（bind 0 后立刻关闭，存在极小竞态）。"""
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind(("127.0.0.1", 0))
        return int(s.getsockname()[1])


def start_auto_server(data_root: Path) -> Tuple[str, int, threading.Event]:
    """
    在独立线程启动 PAEE Server，使用临时 data_root，避免污染正式 data/。
    返回 (host, port, stop_event)；调用方结束时应 stop.set()（daemon 线程随进程退出）。
    """
    from main import load_config, run_server

    host = "127.0.0.1"
    port = free_port()
    # 干净数据目录，防止旧 ser_ver 记录干扰基准
    if data_root.exists():
        shutil.rmtree(data_root)
    data_root.mkdir(parents=True, exist_ok=True)

    cfg = load_config()
    cfg = dict(cfg)
    cfg["storage"] = dict(cfg["storage"])
    cfg["storage"]["local_root"] = str(data_root)
    cfg["network"] = dict(cfg["network"])
    cfg["network"]["host"] = host
    cfg["network"]["port"] = port

    stop = threading.Event()

    def _run() -> None:
        try:
            run_server(cfg)  # 阻塞 accept 循环
        except Exception as exc:  # noqa: BLE001
            if not stop.is_set():
                log.error("auto-server exited: %s", exc)

    threading.Thread(target=_run, daemon=True).start()
    # 轮询直到端口可连
    deadline = time.time() + 8
    while time.time() < deadline:
        try:
            with socket.create_connection((host, port), timeout=0.2):
                return host, port, stop
        except OSError:
            time.sleep(0.05)
    raise SystemExit("auto-server failed to start")


def write_xlsx_summary(
    path: Path,
    phase_rows: List[List[Any]],
    raw: List[Dict[str, Any]],
    *,
    raw_headers: List[str],
    raw_map: Callable[[Dict[str, Any]], List[Any]],
) -> None:
    """
    写两表 Excel：
      汇总 — 分阶段 mean/std
      原始数据 — 每次 trial 明细（由 raw_map 投影列）
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        log.warning("openpyxl missing; skip %s", path)
        return

    headers = [
        "方案",
        "阶段",
        "试验次数",
        "成功次数",
        "成功率(%)",
        "轮数",
        "延迟_mean(ms)",
        "延迟_std(ms)",
        "通信量_mean(bytes)",
        "通信量_std(bytes)",
        "Server 端点",
        "测量说明",
        "时间戳",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"
    ws.append(headers)
    from openpyxl.styles import Alignment, Font

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in phase_rows:
        ws.append(row)

    ws2 = wb.create_sheet("原始数据")
    ws2.append(raw_headers)
    for r in raw:
        ws2.append(raw_map(r))

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
