# -*- coding: utf-8 -*-
"""
experiment/password_length_bench.py
===================================
【实验二】口令长度扫描：固定明文（默认 10MB），长度 8…512。

口径与实验一一致：**不传 π**；τ=H5 含 ct2；ENC_COMMIT 不上传 ct2。
报告 Ext / Enc_proto / Dec_proto 随口令长度的延迟。

用法::
    python -m experiment.password_length_bench --trials 20 --host 35.78.207.231 --port 5202
    python -m experiment.password_length_bench --trials 20 --auto-server
"""

from __future__ import annotations

import argparse
import csv
import logging
import secrets
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from experiment.tcp_benchmark import (  # noqa: E402
    DISPLAY,
    _gen_plaintext,
    run_trials,
)
from experiment.tcp_benchmark_lib import (  # noqa: E402
    check_server,
    mean_std,
    start_auto_server,
)

log = logging.getLogger("paee.pwlen")

DEFAULT_LENGTHS = (8, 16, 32, 64, 128, 256, 512)
DEFAULT_PLAINTEXT_MB = 10
PHASES = (
    ("Ext", "ext_latency_ms", 1),
    ("Enc_proto≈Init", "enc_proto_latency_ms", 2),
    ("Dec_proto≈Rec", "dec_proto_latency_ms", 2),
)


def make_password(length: int, *, alphabet: str = "a") -> str:
    """构造指定字节长度的口令（UTF-8 单字节字符重复）。"""
    if length < 1:
        raise ValueError("password length must be >= 1")
    if len(alphabet.encode("utf-8")) != 1:
        raise ValueError("alphabet must be a single ASCII character")
    return alphabet * length


def parse_lengths(spec: str) -> List[int]:
    """解析 ``8,16,32`` 或 ``8:512``（2 的幂）或 ``8:512:32``（步长）。"""
    spec = spec.strip()
    if not spec:
        return list(DEFAULT_LENGTHS)
    if "," in spec:
        return [int(x) for x in spec.split(",") if x.strip()]
    if ":" in spec:
        parts = [int(x) for x in spec.split(":")]
        if len(parts) == 2:
            lo, hi = parts
            out = []
            n = lo
            while n <= hi:
                out.append(n)
                n *= 2
            if out and out[-1] != hi and hi >= lo:
                # 若 hi 不是 2 的幂，仍包含终点
                if hi not in out:
                    out.append(hi)
            return out
        if len(parts) == 3:
            lo, hi, step = parts
            return list(range(lo, hi + 1, step))
    return [int(spec)]


def summarize_length(
    length: int, rows: List[Dict[str, Any]]
) -> Dict[str, Any]:
    ok = [r for r in rows if r["success"]]
    phases: Dict[str, Dict[str, float]] = {}
    for name, key, rounds in PHASES:
        vals = [float(r[key]) for r in ok]
        m, s = mean_std(vals)
        phases[name] = {
            "rounds": rounds,
            "latency_mean_ms": m,
            "latency_std_ms": s,
        }
    return {
        "password_length": length,
        "n_trials": len(rows),
        "n_success": len(ok),
        "success_rate": (len(ok) / len(rows)) if rows else 0.0,
        "phases": phases,
    }


def _length_header(summaries: List[Dict[str, Any]]) -> List[Any]:
    """表头：阶段 | L=8 | L=16 | …"""
    return ["阶段"] + [f"L={s['password_length']}" for s in summaries]


def _phase_row_mean_std(
    summaries: List[Dict[str, Any]], phase: str
) -> List[Any]:
    """一行：阶段名 + 各长度 mean±std (ms)。"""
    cells: List[Any] = [phase]
    for s in summaries:
        p = s["phases"][phase]
        cells.append(
            f"{p['latency_mean_ms']:.2f} ± {p['latency_std_ms']:.2f}"
        )
    return cells


def _phase_row_numeric(
    summaries: List[Dict[str, Any]], phase: str, *, which: str
) -> List[Any]:
    """一行数值：阶段名 + 各长度 mean 或 std。"""
    key = "latency_mean_ms" if which == "mean" else "latency_std_ms"
    cells: List[Any] = [phase]
    for s in summaries:
        cells.append(round(s["phases"][phase][key], 3))
    return cells


def write_outputs(
    summaries: List[Dict[str, Any]],
    raw: List[Dict[str, Any]],
    out_dir: Path,
    *,
    endpoint: str,
    trials: int,
    timestamp: str,
    plaintext_mb: float = DEFAULT_PLAINTEXT_MB,
) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    note = (
        "【实验二】口令长度扫描；口径同实验一（不传π；轮数=论文箭头）；"
        "口令=单字符重复；Ext 取 Enc 路径内首次 Ext；"
        f"plaintext={plaintext_mb}MB（τ=H5 含 ct2，ENC_COMMIT 不上传 ct2）；"
        f"trials/长度={trials}；Server={endpoint}；"
        "表格式：行=阶段，列=口令长度"
    )
    header = _length_header(summaries)
    sep = "|" + "|".join(["------"] * len(header)) + "|"

    # Markdown：列=口令长度，行=阶段
    md_lines = [
        "# PAEE 口令长度 vs Ext / Enc_proto / Dec_proto 延迟",
        "",
        f"生成时间: {timestamp}",
        note,
        "",
        "## 延迟 mean±std (ms)",
        "",
        "| " + " | ".join(str(h) for h in header) + " |",
        sep,
    ]
    for name, _key, _rounds in PHASES:
        md_lines.append(
            "| " + " | ".join(str(c) for c in _phase_row_mean_std(summaries, name)) + " |"
        )
    md_lines += [
        "",
        "## 成功次数 / 试验次数",
        "",
        "| " + " | ".join(str(h) for h in header) + " |",
        sep,
        "| "
        + " | ".join(
            ["成功/试验"]
            + [f"{s['n_success']}/{s['n_trials']}" for s in summaries]
        )
        + " |",
        "",
    ]
    (out_dir / "password_length_benchmark.md").write_text(
        "\n".join(md_lines), encoding="utf-8"
    )

    # 宽表 CSV（列=长度）
    summary_csv = out_dir / "password_length_benchmark_summary.csv"
    with summary_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        for name, _key, _rounds in PHASES:
            w.writerow(_phase_row_mean_std(summaries, name))
        w.writerow(
            ["成功/试验"] + [f"{s['n_success']}/{s['n_trials']}" for s in summaries]
        )

    # 原始 trial CSV（长表，便于重算）
    if raw:
        csv_path = out_dir / "password_length_benchmark.csv"
        with csv_path.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(raw[0].keys()))
            w.writeheader()
            w.writerows(raw)

    # Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        log.warning("openpyxl missing; skip xlsx")
        return

    wb = Workbook()

    def _style_header(ws) -> None:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    # 汇总：列=口令长度，单元格=mean±std
    ws = wb.active
    ws.title = "汇总"
    ws.append(header)
    _style_header(ws)
    for name, _key, _rounds in PHASES:
        ws.append(_phase_row_mean_std(summaries, name))
    ws.append(
        ["成功/试验"] + [f"{s['n_success']}/{s['n_trials']}" for s in summaries]
    )
    ws.append([])
    ws.append([note])
    ws.append([f"方案={DISPLAY}", f"时间={timestamp}", f"端点={endpoint}"])

    # 分表：纯数值 mean / std（便于画图）
    ws_mean = wb.create_sheet("延迟_mean")
    ws_mean.append(header)
    _style_header(ws_mean)
    for name, _key, _rounds in PHASES:
        ws_mean.append(_phase_row_numeric(summaries, name, which="mean"))

    ws_std = wb.create_sheet("延迟_std")
    ws_std.append(header)
    _style_header(ws_std)
    for name, _key, _rounds in PHASES:
        ws_std.append(_phase_row_numeric(summaries, name, which="std"))

    ws3 = wb.create_sheet("原始数据")
    raw_headers = [
        "口令长度",
        "trial_id",
        "id",
        "success",
        "Ext延迟(ms)",
        "Enc_proto延迟(ms)",
        "Dec_proto延迟(ms)",
        "Ext通信(B)",
        "Enc_proto通信(B)",
        "Dec_proto通信(B)",
        "error",
    ]
    ws3.append(raw_headers)
    _style_header(ws3)
    for r in raw:
        ws3.append(
            [
                r["password_length"],
                r["trial_id"],
                r["id"],
                r["success"],
                round(r["ext_latency_ms"], 3),
                round(r["enc_proto_latency_ms"], 3),
                round(r["dec_proto_latency_ms"], 3),
                r["ext_comm_bytes"],
                r["enc_proto_comm_bytes"],
                r["dec_proto_comm_bytes"],
                r["error"],
            ]
        )

    path = out_dir / "password_length_benchmark.xlsx"
    wb.save(path)
    log.info("wrote %s", path)


def run_sweep(
    lengths: Sequence[int],
    *,
    trials: int,
    host: str,
    port: int,
    lambda_bytes: int,
    id_prefix: str,
    alphabet: str,
    plaintext: bytes,
) -> tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    summaries: List[Dict[str, Any]] = []
    all_raw: List[Dict[str, Any]] = []

    for L in lengths:
        pw = make_password(L, alphabet=alphabet)
        assert len(pw.encode("utf-8")) == L
        log.info("===== password length=%d |pt|=%d =====", L, len(plaintext))
        rows = run_trials(
            trials,
            host=host,
            port=port,
            password=pw,
            id_prefix=f"{id_prefix}_L{L}",
            lambda_bytes=lambda_bytes,
            plaintext=plaintext,
        )
        for r in rows:
            r["password_length"] = L
        summaries.append(summarize_length(L, rows))
        all_raw.extend(rows)
        s = summaries[-1]
        log.info(
            "L=%d Ext=%.2f±%.2f Enc=%.2f±%.2f Dec=%.2f±%.2f (%d/%d ok)",
            L,
            s["phases"]["Ext"]["latency_mean_ms"],
            s["phases"]["Ext"]["latency_std_ms"],
            s["phases"]["Enc_proto≈Init"]["latency_mean_ms"],
            s["phases"]["Enc_proto≈Init"]["latency_std_ms"],
            s["phases"]["Dec_proto≈Rec"]["latency_mean_ms"],
            s["phases"]["Dec_proto≈Rec"]["latency_std_ms"],
            s["n_success"],
            s["n_trials"],
        )
    return summaries, all_raw


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="PAEE password-length latency sweep (Ext / Enc_proto / Dec_proto)"
    )
    parser.add_argument("--trials", type=int, default=20, help="每个口令长度的重复次数")
    parser.add_argument(
        "--lengths",
        default="8,16,32,64,128,256,512",
        help="口令长度列表，或 8:512（2的幂），或 8:512:32（步长）",
    )
    parser.add_argument("--host", default=None, help="Server 地址（默认读 config.yaml）")
    parser.add_argument("--port", type=int, default=None)
    parser.add_argument("--auto-server", action="store_true")
    parser.add_argument("--alphabet", default="a", help="口令填充用的单 ASCII 字符")
    parser.add_argument(
        "--plaintext-mb",
        type=float,
        default=DEFAULT_PLAINTEXT_MB,
        help=f"Enc/Dec 明文大小 MB（默认 {DEFAULT_PLAINTEXT_MB}；0=空明文）",
    )
    parser.add_argument(
        "--id-prefix",
        default="pwlen",
        help="用户 id 前缀；每次运行会自动追加时间戳后缀，避免 ID_EXISTS",
    )
    parser.add_argument(
        "--out-dir",
        default=str(ROOT / "experiment" / "output"),
    )
    parser.add_argument("-q", "--quiet", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )

    from main import load_config

    cfg = load_config()
    lambda_bytes = int(cfg["crypto"]["lambda_bytes"])
    lengths = parse_lengths(args.lengths)
    if any(L < 1 for L in lengths):
        raise SystemExit("password lengths must be >= 1")

    run_tag = datetime.now().strftime("%Y%m%d%H%M%S") + secrets.token_hex(2)
    id_prefix = f"{args.id_prefix}_{run_tag}"
    plaintext = _gen_plaintext(args.plaintext_mb)
    log.info(
        "id_prefix=%s plaintext_mb=%s |pt|=%d",
        id_prefix,
        args.plaintext_mb,
        len(plaintext),
    )

    stop = None
    host = args.host if args.host is not None else str(cfg["network"]["host"])
    port = args.port if args.port is not None else int(cfg["network"]["port"])
    if host in ("0.0.0.0", "::"):
        host = "127.0.0.1"

    if args.auto_server:
        host, port, stop = start_auto_server(ROOT / "data" / "bench_pwlen_server")
    else:
        check_server(host, port)

    try:
        summaries, raw = run_sweep(
            lengths,
            trials=args.trials,
            host=host,
            port=port,
            lambda_bytes=lambda_bytes,
            id_prefix=id_prefix,
            alphabet=args.alphabet,
            plaintext=plaintext,
        )
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        write_outputs(
            summaries,
            raw,
            Path(args.out_dir),
            endpoint=f"{host}:{port}",
            trials=args.trials,
            timestamp=ts,
            plaintext_mb=args.plaintext_mb,
        )
        print()
        header = _length_header(summaries)
        col_w = max(14, max(len(str(h)) for h in header[1:]) + 2)
        print(f"{header[0]:<18}" + "".join(f"{h:>{col_w}}" for h in header[1:]))
        for name, _key, _rounds in PHASES:
            row = _phase_row_mean_std(summaries, name)
            print(f"{row[0]:<18}" + "".join(f"{c:>{col_w}}" for c in row[1:]))
        print(f"\n结果目录: {args.out_dir}")
        print("  password_length_benchmark.xlsx / .md / _summary.csv / .csv")
    finally:
        if stop is not None:
            stop.set()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
